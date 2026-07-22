"""Pure unit tests for the CP file parsers — in-memory fixtures, no DB/network.

Timesheet fixtures mirror the BusyBusy export headers verbatim (including the
double space in "Subproject 1  #"); the Gusto fixture mirrors the 65-column
payroll-detail layout with its preamble rows and trailing Totals row.
"""

import io
from datetime import date, datetime
from decimal import Decimal

import pytest
from openpyxl import Workbook

from app.services.payroll_detail_parser import COLUMN_MAP, parse_payroll_detail
from app.services.payroll_timesheet_parser import (
    get_week_dates,
    parse_csv_timesheet,
    parse_datetime,
    parse_excel_timesheet,
    parse_hours_string,
    parse_time_string,
    parse_timesheet,
)

TIMESHEET_HEADERS = [
    "Employee ID",
    "First Name",
    "Last Name",
    "Start",
    "End",
    "Break Total",
    "Total",
    "Customer",
    "Project #",
    "Project",
    "Subproject 1  #",  # double space is verbatim from the export
    "Subproject 1",
    "Cost Code",
    "Cost Code Desc.",
    "Description",
]

CSV_CONTENT = (
    ",".join(TIMESHEET_HEADERS)
    + "\n"
    + "E100,Juan,Perez,2026-01-26T05:03:00.000-08:00,2026-01-26T13:33:00.000-08:00,"
    + "00:30,08:00,LVCVA,6370,6370 - Terminal 3 Remodel,01,Rough,26-05,Conduit,Pulled feeders\n"
    + ",,,,,,,,,,,,,,\n"  # fully blank row -> skipped
    + "E101,Maria,Lopez,2026-01-26T06:00:00.000-08:00,2026-01-26T16:30:00.000-08:00,"
    + "00:30,10:30,,6371,Fab Shop,,,,,\n"
    + "E102,No,Start,,2026-01-26T16:00:00.000-08:00,00:00,08:00,,,,,,,,\n"  # no Start -> skipped
)


# ── low-level helpers ───────────────────────────────────────────────────────


def test_parse_time_string():
    assert parse_time_string("00:30") == 30
    assert parse_time_string("01:15") == 75
    assert parse_time_string("") == 0
    assert parse_time_string("garbage") == 0


def test_parse_hours_string():
    assert parse_hours_string("08:30") == Decimal("8.5")
    assert parse_hours_string("10:00") == Decimal("10")
    assert parse_hours_string("") == Decimal("0.00")
    assert parse_hours_string("nope") == Decimal("0.00")


def test_parse_datetime_strips_timezone():
    dt = parse_datetime("2026-01-26T05:03:00.000-08:00")
    assert dt == datetime(2026, 1, 26, 5, 3)
    assert dt.tzinfo is None
    assert parse_datetime("2026-01-26T05:03:00Z") == datetime(2026, 1, 26, 5, 3)
    assert parse_datetime("") is None
    assert parse_datetime("not-a-date") is None


def test_get_week_dates_snaps_to_sunday_saturday():
    # 2026-07-15 is a Wednesday
    assert get_week_dates(date(2026, 7, 15)) == (date(2026, 7, 12), date(2026, 7, 18))
    # A Sunday is its own week start
    assert get_week_dates(date(2026, 7, 12)) == (date(2026, 7, 12), date(2026, 7, 18))
    # A Saturday snaps back to the preceding Sunday
    assert get_week_dates(date(2026, 7, 18)) == (date(2026, 7, 12), date(2026, 7, 18))


# ── CSV timesheet ───────────────────────────────────────────────────────────


def test_parse_csv_timesheet_rows():
    entries = parse_csv_timesheet(CSV_CONTENT.encode("utf-8-sig"))  # BOM-prefixed
    assert len(entries) == 2  # blank row and missing-Start row are skipped

    e = entries[0]
    assert (e.employee_id, e.first_name, e.last_name) == ("E100", "Juan", "Perez")
    assert e.work_date == date(2026, 1, 26)
    assert e.start_time == datetime(2026, 1, 26, 5, 3)
    assert e.end_time == datetime(2026, 1, 26, 13, 33)
    assert e.start_time.tzinfo is None and e.end_time.tzinfo is None
    assert e.break_total_minutes == 30
    assert e.total_hours == Decimal("8")
    assert e.customer == "LVCVA"
    assert e.project_number == "6370"
    assert e.project_name == "6370 - Terminal 3 Remodel"
    assert e.subproject_1_number == "01"
    assert e.subproject_1_name == "Rough"
    assert e.cost_code == "26-05"
    assert e.cost_code_desc == "Conduit"
    assert e.description == "Pulled feeders"

    e2 = entries[1]
    assert e2.first_name == "Maria"
    assert e2.total_hours == Decimal("10.5")
    assert e2.customer is None  # empty cells become None
    assert e2.subproject_1_number is None
    assert e2.project_number == "6371"


# ── Excel timesheet ─────────────────────────────────────────────────────────


def _timesheet_xlsx() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.append(TIMESHEET_HEADERS)
    ws.append(
        [
            "E100",
            "Juan",
            "Perez",
            datetime(2026, 1, 26, 5, 0),
            datetime(2026, 1, 26, 13, 30),
            "00:30",
            "08:00",
            "LVCVA",
            "6370",
            "6370 - Terminal 3 Remodel",
            "01",  # stays a string because the column below is non-numeric
            "Rough",
            "26-05",
            "Conduit",
            "Pulled feeders",
        ]
    )
    ws.append([None] * len(TIMESHEET_HEADERS))  # blank row -> skipped
    ws.append(
        [
            None,
            "Maria",
            "Lopez",
            "2026-01-26T06:00:00.000-08:00",  # string datetimes also accepted
            "2026-01-26T16:30:00.000-08:00",
            None,
            "10:30",
            None,
            "6371 - Fab Shop",  # keeps the Project #/Subproject columns object-dtype
            "Fab Shop",
            "02B",
            None,
            None,
            None,
            None,
        ]
    )
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_parse_excel_timesheet_rows():
    entries = parse_excel_timesheet(_timesheet_xlsx())
    assert len(entries) == 2

    e = entries[0]
    assert (e.employee_id, e.first_name, e.last_name) == ("E100", "Juan", "Perez")
    assert e.work_date == date(2026, 1, 26)
    assert e.start_time == datetime(2026, 1, 26, 5, 0)
    assert e.end_time == datetime(2026, 1, 26, 13, 30)
    assert e.start_time.tzinfo is None
    assert e.break_total_minutes == 30
    assert e.total_hours == Decimal("8")
    assert e.project_number == "6370"
    assert e.subproject_1_number == "01"

    e2 = entries[1]
    assert e2.employee_id is None
    assert e2.start_time == datetime(2026, 1, 26, 6, 0)
    assert e2.end_time == datetime(2026, 1, 26, 16, 30)
    assert e2.total_hours == Decimal("10.5")
    assert e2.break_total_minutes == 0
    assert e2.customer is None
    assert e2.project_number == "6371 - Fab Shop"
    assert e2.project_name == "Fab Shop"
    assert e2.subproject_1_number == "02B"


def test_parse_excel_numeric_project_column_arrives_as_float_string():
    # pandas type inference: a purely numeric "Project #" column (plus blanks)
    # becomes float64, so 6370 arrives as "6370.0". Ingest matching must
    # normalize numeric-looking values rather than compare verbatim.
    wb = Workbook()
    ws = wb.active
    ws.append(TIMESHEET_HEADERS)
    ws.append(
        [
            "E100",
            "Juan",
            "Perez",
            datetime(2026, 1, 26, 5, 0),
            datetime(2026, 1, 26, 13, 30),
            "00:30",
            "08:00",
            None,
            6370,
            "Terminal 3",
            None,
            None,
            None,
            None,
            None,
        ]
    )
    ws.append(
        [
            "E101",
            "Maria",
            "Lopez",
            datetime(2026, 1, 26, 6, 0),
            datetime(2026, 1, 26, 14, 0),
            "00:30",
            "07:30",
            None,
            None,  # blank cell forces the numeric column to float64
            None,
            None,
            None,
            None,
            None,
            None,
        ]
    )
    buf = io.BytesIO()
    wb.save(buf)
    entries = parse_excel_timesheet(buf.getvalue())
    assert entries[0].project_number == "6370.0"
    assert entries[1].project_number is None


def test_parse_timesheet_dispatch():
    assert len(parse_timesheet(CSV_CONTENT.encode("utf-8"), "week 27.CSV")) == 2
    assert len(parse_timesheet(_timesheet_xlsx(), "week 27.xlsx")) == 2
    with pytest.raises(ValueError, match="Unsupported file format"):
        parse_timesheet(b"", "week27.pdf")


# ── Gusto payroll detail ────────────────────────────────────────────────────

N_COLS = 65


def _detail_row(name, pay_date, period, **fields) -> list:
    """Build a 65-cell row, placing values by COLUMN_MAP field name."""
    row = [None] * N_COLS
    row[0], row[1], row[2] = name, pay_date, period
    index_of = {field: idx for idx, field in COLUMN_MAP.items()}
    for field, value in fields.items():
        row[index_of[field]] = value
    return row


def _detail_xlsx() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.append(["G3 Electrical"] + [None] * (N_COLS - 1))  # row 0: company
    ws.append(["Payroll Detail"] + [None] * (N_COLS - 1))  # row 1: title
    ws.append([None] * N_COLS)  # row 2: blank
    ws.append(["06/28/2026 - 07/04/2026"] + [None] * (N_COLS - 1))  # row 3: range
    ws.append(["Name"] + [f"Col {i}" for i in range(1, N_COLS)])    # row 4: headers
    ws.append(
        _detail_row(
            "Perez, Juan",
            "07/10/2026",
            "06/28/2026 - 07/04/2026",
            hours_total=44,
            hours_regular=40,
            hours_ot=4,
            gross_pay_total=2130.5,
            gross_pay_regular=1720,
            gross_pay_ot=258,
            other_pay_qot=12.346,  # quantized to 0.01
            employee_taxes_total=402.17,
            net_pay=1728.33,
            employer_taxes_total=190.02,
            total_payroll_cost=2320.52,
        )
    )
    ws.append(_detail_row("Lopez, Maria", "07/10/2026", None, hours_total=32, net_pay=990))
    ws.append([None] * N_COLS)  # blank row -> skipped
    ws.append(_detail_row("Totals", None, None, hours_total=76))  # totals row -> skipped
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_parse_payroll_detail():
    entries = parse_payroll_detail(_detail_xlsx(), "payroll_detail.xlsx")
    assert len(entries) == 2  # blank + Totals rows skipped

    e = entries[0]
    assert e.employee_name == "Perez, Juan"
    assert e.pay_date == date(2026, 7, 10)
    assert e.time_period == "06/28/2026 - 07/04/2026"
    assert e.hours_total == Decimal("44.00")
    assert e.hours_regular == Decimal("40.00")
    assert e.hours_ot == Decimal("4.00")
    assert e.hours_holiday is None  # empty cell -> None
    assert e.gross_pay_total == Decimal("2130.50")
    assert e.gross_pay_ot == Decimal("258.00")
    assert e.other_pay_qot == Decimal("12.35")  # quantized
    assert e.employee_taxes_total == Decimal("402.17")
    assert e.net_pay == Decimal("1728.33")
    assert e.employer_taxes_total == Decimal("190.02")
    assert e.total_payroll_cost == Decimal("2320.52")

    e2 = entries[1]
    assert e2.employee_name == "Lopez, Maria"
    assert e2.time_period is None
    assert e2.hours_total == Decimal("32.00")
    assert e2.net_pay == Decimal("990.00")


def test_parse_payroll_detail_missing_header_raises():
    wb = Workbook()
    ws = wb.active
    ws.append(["Just some sheet"])
    ws.append(["with no Name header"])
    buf = io.BytesIO()
    wb.save(buf)
    with pytest.raises(ValueError, match="header row"):
        parse_payroll_detail(buf.getvalue(), "bad.xlsx")


# ── Regression tests for the review fixes ───────────────────────────────────


def test_parse_hours_string_accepts_plain_decimal():
    # An export whose Total column is numeric (not H:MM) must not silently zero
    # every worked hour.
    assert parse_hours_string("10") == Decimal("10")
    assert parse_hours_string("10.5") == Decimal("10.5")
    assert parse_hours_string(" 8.25 ") == Decimal("8.25")
    # Genuinely unparseable still yields 0.00, and H:MM still works.
    assert parse_hours_string("nope") == Decimal("0.00")
    assert parse_hours_string("08:30") == Decimal("8.5")


def test_parse_csv_timesheet_short_row_does_not_crash():
    # A data row with fewer cells than the header makes csv.DictReader fill the
    # rest with None; the parser must not raise AttributeError on None.strip().
    content = (
        ",".join(TIMESHEET_HEADERS)
        + "\n"
        + "E100,Juan,Perez,2026-01-26T05:03:00-08:00,2026-01-26T13:33:00-08:00,00:30,08:00\n"
    )
    entries = parse_csv_timesheet(content.encode("utf-8"))
    assert len(entries) == 1
    assert entries[0].customer is None and entries[0].description is None


def test_parse_timesheet_corrupt_excel_raises_valueerror():
    # A truncated/corrupt xlsx raises BadZipFile inside pandas — the parser must
    # normalize it to ValueError so the upload handler returns 400, not 500.
    with pytest.raises(ValueError):
        parse_timesheet(b"PK\x03\x04 not really a zip", "week.xlsx")


def test_parse_payroll_detail_wrong_column_count_raises():
    # A drifted Gusto layout (not 65 columns) must be refused rather than parsed
    # positionally into the wrong money fields.
    wb = Workbook()
    ws = wb.active
    ws.append(["G3 Electrical"] + [None] * 63)
    ws.append(["Payroll Detail"] + [None] * 63)
    ws.append([None] * 64)
    ws.append(["06/28/2026 - 07/04/2026"] + [None] * 63)
    ws.append(["Name"] + [f"Col {i}" for i in range(1, 64)])  # 64 columns, not 65
    ws.append(["Perez, Juan"] + [None] * 63)
    buf = io.BytesIO()
    wb.save(buf)
    with pytest.raises(ValueError, match="layout"):
        parse_payroll_detail(buf.getvalue(), "detail.xlsx")


def test_parse_pay_date_narrows_datetime_to_date():
    from app.services.payroll_detail_parser import _parse_pay_date

    assert _parse_pay_date(datetime(2026, 7, 10, 9, 30)) == date(2026, 7, 10)
    assert _parse_pay_date(date(2026, 7, 10)) == date(2026, 7, 10)
    assert _parse_pay_date("07/10/2026") == date(2026, 7, 10)


def test_weekly_summaries_drop_out_of_week_entries():
    from app.services.payroll_ot import (
        TimeEntryInput,
        calculate_weekly_summaries,
        out_of_week_entries,
    )

    week_start = date(2026, 7, 12)  # Sunday
    in_week = TimeEntryInput(
        employee_id="e1", project_id="p1", work_date=date(2026, 7, 12),
        start_time=datetime(2026, 7, 12, 7), total_hours=Decimal("8"), shift_type="regular",
    )
    next_week = TimeEntryInput(
        employee_id="e1", project_id="p1", work_date=date(2026, 7, 19),  # next Sunday
        start_time=datetime(2026, 7, 19, 7), total_hours=Decimal("8"), shift_type="regular",
    )
    summaries = calculate_weekly_summaries([in_week, next_week], week_start)
    [summary] = summaries
    # Only the in-week 8h lands in Sunday's ST column — the next-week entry does
    # NOT fold into the same weekday column (which would have shown 16h).
    assert summary.daily_st[0] == Decimal("8")
    assert summary.total_hours == Decimal("8")
    # And the dropped entry is surfaced for the caller.
    assert out_of_week_entries([in_week, next_week], week_start) == [next_week]
