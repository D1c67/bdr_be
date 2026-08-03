"""Bid Invitations .xlsx export — mirrors assets/Bid Invitations Template.xlsx:
weekday-coloured bid/needs-by dates (Mon yellow → Fri green), the win/loss tally
in the banner, quotes x/y green once everything is back, one row per bid.
Follows rfq_excel.py's build-in-memory / return-bytes shape. The workbook is
pre-set to print on letter landscape, fit to page width."""

import io
from datetime import datetime, timezone

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.properties import PageSetupProperties

from app.services.bid_invitations import REPORT_TZ, _parse
from app.services.cpr_generation import _formula_safe

# Weekday fills lifted from the template (Mon–Fri legend; weekends unfilled).
_WEEKDAY_FILLS = {
    0: "FFFF99",  # Monday — yellow
    1: "CCFFFF",  # Tuesday — blue
    2: "FFCCFF",  # Wednesday — pink
    3: "FFC000",  # Thursday — orange
    4: "CCFF33",  # Friday — green
}

_NAVY = "202159"  # G3 brand
_QUOTES_DONE_FILL = PatternFill("solid", fgColor="C6EFCE")
_QUOTES_DONE_FONT = Font(color="006100")

_STATUS_LABELS = {
    "won": "Win",
    "lost": "Loss",
    "no_award": "No Award",
    "declined": "No Bid",
    "abandoned": "No Bid (abandoned)",
    "sent": "Waiting on results",
    "active": "In progress",
}

_RANGE_LABELS = {
    "today": "Today & upcoming",
    "wtd": "Week to date & upcoming",
    "mtd": "Month to date & upcoming",
    "ytd": "Year to date & upcoming",
    "yesterday": "Yesterday",
    "last_week": "Last week",
    "last_month": "Last month",
    "last_year": "Last year",
    "past_5_years": "Past 5 years",
    "custom": "Custom range",
}


def _local(ts: str | None) -> datetime | None:
    dt = _parse(ts)
    return dt.astimezone(REPORT_TZ) if dt else None


def _fmt_date(ts: str | None) -> str:
    dt = _local(ts)
    return dt.strftime("%-m/%-d/%Y") if dt else ""


def _fmt_time(ts: str | None) -> str:
    dt = _local(ts)
    return dt.strftime("%-I:%M%p").lower() if dt else ""


def _fmt_datetime(ts: str | None) -> str:
    dt = _local(ts)
    return f"{dt.strftime('%-m/%-d/%Y')}  {dt.strftime('%-I:%M%p').lower()}" if dt else ""


def _fmt_day(d: str | None) -> str:
    if not d:
        return ""
    y, m, dd = d[:10].split("-")
    return f"{int(m)}/{int(dd)}/{y}"


def _weekday_fill(ts_or_date: str | None) -> PatternFill | None:
    if not ts_or_date:
        return None
    if len(ts_or_date) > 10:  # timestamp — weekday in report tz
        dt = _local(ts_or_date)
        wd = dt.weekday() if dt else None
    else:  # bare date
        y, m, d = ts_or_date[:10].split("-")
        wd = datetime(int(y), int(m), int(d)).weekday()
    color = _WEEKDAY_FILLS.get(wd) if wd is not None else None
    return PatternFill("solid", fgColor=color) if color else None


def _gc_lines(gcs: list[dict], key: str, fmt) -> str:
    """One line per GC that has the value, prefixed with the GC name when the
    project has several GCs — matches how the sheet stacks multi-GC info."""
    vals = [(g["name"], g.get(key)) for g in gcs if g.get(key)]
    if not vals:
        return ""
    if len(gcs) == 1:
        return fmt(vals[0][1])
    return "\n".join(f"{fmt(v)} — {name}" for name, v in vals)


def build_bid_invitations_workbook(payload: dict) -> bytes:
    meta, summary, rows = payload["meta"], payload["summary"], payload["rows"]
    show_bid = meta["bid_dates_visible"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Bid Invitations"

    headers = (["Bid Date", "Bid Time"] if show_bid else []) + [
        "Project", "GC Needs By", "Invitation to Bid", "Est. Start & Finish",
        "GC", "Received from Estimator", "Quotes", "Win / Loss",
        "Date Sent to GC", "Time Sent to GC", "Bid Late / Early",
    ]
    ncols = len(headers)
    widths = ([12, 9] if show_bid else []) + [30, 16, 13, 22, 22, 14, 9, 13, 14, 12, 11]
    for i, wd in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = wd

    # Row 1 — title banner.
    now_local = datetime.now(timezone.utc).astimezone(REPORT_TZ)
    range_label = _RANGE_LABELS.get(meta["range"], meta["range"])
    window = _fmt_date(meta["date_from"]) + " – " + (
        _fmt_date(meta["date_to"]) if meta["date_to"] else "open"
    )
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(row=1, column=1,
                value=f"Bid Invitations — {range_label} ({window})  ·  "
                      f"generated {now_local.strftime('%-m/%-d/%Y %-I:%M%p').lower()}")
    c.font = Font(bold=True, size=13, color=_NAVY)

    # Row 2 — weekday legend + win/loss tally (like the sheet's banner row).
    half = max(1, ncols // 2)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=half)
    ws.cell(row=2, column=1,
            value="Mon yellow · Tue blue · Wed pink · Thu orange · Fri green")
    ws.merge_cells(start_row=2, start_column=half + 1, end_row=2, end_column=ncols)
    t = ws.cell(
        row=2, column=half + 1,
        value=(f"Bids: {summary['bids']}  ·  Win: {summary['won']}  ·  "
               f"Loss: {summary['lost']}  ·  No Award: {summary['no_award']}  ·  "
               f"No Bid: {summary['no_bid']}  ·  Waiting: {summary['waiting']}  ·  "
               f"In progress: {summary['active']}"),
    )
    t.font = Font(bold=True)
    t.alignment = Alignment(horizontal="right")

    # Row 3 — column headers.
    header_fill = PatternFill("solid", fgColor=_NAVY)
    header_font = Font(bold=True, color="FFFFFF")
    thin = Side(style="thin", color="B7B7C4")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=i, value=h)
        cell.fill, cell.font, cell.border = header_fill, header_font, border
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")

    wrap = Alignment(wrap_text=True, vertical="top")
    for rix, row in enumerate(rows, start=4):
        gcs = row["gcs"]
        est = " – ".join(v for v in (_fmt_day(row.get("est_start_date")),
                                     _fmt_day(row.get("est_finish_date"))) if v)
        notes = "\n".join(v for v in (row.get("wage_label"), row.get("labor_note")) if v)
        quotes = (f"{row['quotes_received']}/{row['quotes_requested']}"
                  if row["quotes_requested"] or row["quotes_received"] else "")
        on_time = {"on_time": "On time", "late": "Late"}.get(row["on_time"], "")

        values = ([_fmt_date(row.get("bid_at")), _fmt_time(row.get("bid_at"))]
                  if show_bid else []) + [
            f"{row['number']} — {row['name']}" if row.get("number") else row["name"],
            _gc_lines(gcs, "needs_by", _fmt_day),
            _fmt_date(row.get("invitation_at")),
            "\n".join(v for v in (est, notes) if v),
            " / ".join(g["name"] for g in gcs),
            _fmt_datetime(row.get("estimator_returned_at")),
            quotes,
            _STATUS_LABELS.get(row["status"], row["status"]),
            _gc_lines(gcs, "sent_at", _fmt_date),
            _gc_lines(gcs, "sent_at", _fmt_time),
            on_time,
        ]
        for cix, v in enumerate(values, start=1):
            # _formula_safe: project names, GC names and labor notes are
            # free text — a leading '=' would otherwise land as a live
            # formula cell (CWE-1236), same guard as the CPR exports.
            cell = ws.cell(row=rix, column=cix, value=_formula_safe(v))
            cell.border = border
            cell.alignment = wrap

        col = 1
        if show_bid:
            fill = _weekday_fill(row.get("bid_at"))
            if fill:
                ws.cell(row=rix, column=1).fill = fill
                ws.cell(row=rix, column=2).fill = fill
            col = 3
        needs = {g.get("needs_by") for g in gcs if g.get("needs_by")}
        if len(needs) == 1:  # a single shared date gets its weekday colour
            fill = _weekday_fill(next(iter(needs)))
            if fill:
                ws.cell(row=rix, column=col + 1).fill = fill
        if row["quotes_requested"] and row["quotes_received"] >= row["quotes_requested"]:
            qcell = ws.cell(row=rix, column=col + 6)
            qcell.fill, qcell.font = _QUOTES_DONE_FILL, _QUOTES_DONE_FONT

    # Print set-up: letter landscape, fit to width, headers repeat on each page.
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.print_title_rows = "3:3"
    ws.page_margins.left = ws.page_margins.right = 0.4
    ws.page_margins.top = ws.page_margins.bottom = 0.5
    ws.freeze_panes = "A4"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
