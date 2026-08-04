"""Generate per-category RFQ Excel files from confirmed line items.

The document is the vendor-facing face of G3 — it goes out attached to the RFQ
email (converted to an immutable PDF first, see rfq_sending._as_immutable_pdf),
so it is laid out as branded letterhead rather than a bare grid, using the same
theme as the outbound email shell (see email_branding): the G3 logo, navy
#202159, a red hairline rule, and a navy banner over a bordered table.

Structure:

    [logo]                                   REQUEST FOR QUOTE
                                             6954 - Sunset Ridge
                                             Issued August 3, 2026
    ───────────────────────── red hairline ─────────────────────────
    █ LIGHTING                       QUOTES DUE  FRIDAY, AUGUST 8TH █
      SR.NO │ DESCRIPTION │ QUANTITY │ UNIT │ NOTES
      …one row per material…
      instructions + G3 contact footer

The columns stay the BOQ's four (plus NOTES) and there is still no price
column — the vendor quotes off their own sheet.
"""

import io
from datetime import datetime
from typing import Any
from zoneinfo import ZoneInfo

from app.core.config import get_settings
from app.services import email_branding
from app.services.cpr_generation import _formula_safe
from app.services.datetime_format import format_bid_datetime

_HEADERS = ["SR.NO", "DESCRIPTION", "QUANTITY", "UNIT", "NOTES"]
_WIDTHS = [9, 58, 12, 10, 30]

# Theme — the email shell's palette, as ARGB-less hex for openpyxl.
_NAVY = "202159"
_RED = "951E2D"
_BORDER = "D8DBE0"
_MUTED = "6A6F78"
_BAND = "F5F6F9"  # alternating row tint

_FOOTER_NOTE = (
    "Please quote the items above. Quantities are provided for quoting reference; "
    "verify against the drawings and specifications, and note any substitutions, "
    "exclusions or items you are unable to quote with your pricing."
)

# Letterhead geometry: the logo floats over rows 1–4 (it is wider than the
# SR.NO column and spills into DESCRIPTION, which is empty up there). Row
# heights are in points (1pt = 4/3 px), so 4 x 17pt ~= 91px of vertical room
# for an 82px logo inset 8px/5px from the corner.
_LOGO_ROWS = 4
_LOGO_ROW_PT = 17
_LOGO_PX = 82
_LOGO_INSET_PX = (8, 5)  # x, y


def _issued_today() -> str:
    tz = ZoneInfo(get_settings().display_timezone)
    d = datetime.now(tz)
    return f"{d:%B} {d.day}, {d:%Y}"


def _project_label(project: dict[str, Any] | None) -> str:
    if not project:
        return ""
    number, name = project.get("number"), project.get("name") or ""
    return f"{number} - {name}" if number else name


def build_rfq_workbook(
    category_name: str,
    line_items: list[dict[str, Any]],
    project: dict[str, Any] | None = None,
) -> bytes:
    """Build an .xlsx for one RFQ category and return its bytes.

    `project` is the projects row the RFQ belongs to — its number/name and
    `due_from_vendors_at` populate the letterhead. It is optional so callers
    without project context still get a valid (unpersonalized) sheet. The due
    date is a snapshot taken when the workbook is generated; the RFQ email body
    carries the authoritative one at send time.
    """
    import openpyxl
    from openpyxl.drawing.image import Image as XlsxImage
    from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
    from openpyxl.drawing.xdr import XDRPositiveSize2D
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.utils.units import pixels_to_EMU
    from openpyxl.worksheet.properties import PageSetupProperties

    ncols = len(_HEADERS)
    last_col = get_column_letter(ncols)
    category = (category_name or "RFQ").strip()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = (category or "RFQ")[:31]  # Excel caps sheet names at 31 chars
    ws.sheet_view.showGridLines = False  # the table's own borders are the grid

    wb.properties.creator = "G3 Electrical"
    wb.properties.title = f"Request for Quote - {category}"

    for i, width in enumerate(_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # ---- Letterhead: logo left, document title + project right ----------
    for r in range(1, _LOGO_ROWS + 1):
        ws.row_dimensions[r].height = _LOGO_ROW_PT

    logo = XlsxImage(io.BytesIO(email_branding.logo_bytes()))
    aspect = logo.width / logo.height  # capture before either is overwritten
    logo.height, logo.width = _LOGO_PX, round(_LOGO_PX * aspect)
    # A plain "A1" anchor would jam the logo into the very corner of the page;
    # a OneCellAnchor lets it sit inset like printed letterhead.
    dx, dy = _LOGO_INSET_PX
    logo.anchor = OneCellAnchor(
        _from=AnchorMarker(col=0, colOff=pixels_to_EMU(dx), row=0, rowOff=pixels_to_EMU(dy)),
        ext=XDRPositiveSize2D(pixels_to_EMU(logo.width), pixels_to_EMU(logo.height)),
    )
    ws.add_image(logo)

    def _headline(row: int, text: str, font: Font) -> None:
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=ncols)
        cell = ws.cell(row=row, column=2, value=text)
        cell.font = font
        cell.alignment = Alignment(horizontal="right", vertical="center")

    _headline(1, "REQUEST FOR QUOTE", Font(bold=True, size=17, color=_NAVY))
    _headline(2, _project_label(project), Font(size=11, color=_NAVY))
    _headline(3, f"Issued {_issued_today()}", Font(size=9, color=_MUTED))
    _headline(4, "G3 Electrical  ·  Office (702) 916-3355", Font(size=9, color=_MUTED))

    # Red hairline under the letterhead — the email shell's 3px accent rule.
    ws.row_dimensions[5].height = 3.5
    for c in range(1, ncols + 1):
        ws.cell(row=5, column=c).fill = PatternFill("solid", fgColor=_RED)

    ws.row_dimensions[6].height = 7  # breathing room

    # ---- Navy banner: category on the left, vendor due date on the right --
    banner_row = 7
    ws.row_dimensions[banner_row].height = 26
    navy_fill = PatternFill("solid", fgColor=_NAVY)
    for c in range(1, ncols + 1):
        ws.cell(row=banner_row, column=c).fill = navy_fill

    ws.merge_cells(start_row=banner_row, start_column=1, end_row=banner_row, end_column=3)
    left = ws.cell(row=banner_row, column=1, value=category.upper())
    left.font = Font(bold=True, size=13, color="FFFFFF")
    left.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    due = project.get("due_from_vendors_at") if project else None
    ws.merge_cells(start_row=banner_row, start_column=4, end_row=banner_row, end_column=ncols)
    right = ws.cell(
        row=banner_row,
        column=4,
        value=f"QUOTES DUE  {format_bid_datetime(due).upper()}" if due else "",
    )
    right.font = Font(bold=True, size=9, color="FFFFFF")
    right.alignment = Alignment(horizontal="right", vertical="center", indent=1)

    # ---- Table --------------------------------------------------------
    thin = Side(style="thin", color=_BORDER)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    header_row = banner_row + 1
    ws.row_dimensions[header_row].height = 20
    header_fill = PatternFill("solid", fgColor=_NAVY)
    for col, name in enumerate(_HEADERS, start=1):
        c = ws.cell(row=header_row, column=col, value=name)
        c.font = Font(bold=True, size=10, color="FFFFFF")
        c.fill = header_fill
        c.border = border
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Per-column alignment: numbers right, codes/units centered, text wrapped.
    # The indents are cell padding — without them text sits on the gridline.
    aligns = [
        Alignment(horizontal="center", vertical="top"),
        Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1),
        Alignment(horizontal="right", vertical="top", indent=1),
        Alignment(horizontal="center", vertical="top"),
        Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1),
    ]
    band_fill = PatternFill("solid", fgColor=_BAND)
    body_font = Font(size=10)

    r = header_row + 1
    for i, item in enumerate(line_items):
        qty = item.get("quantity")
        values = [
            item.get("sr_no"),
            item.get("description"),
            float(qty) if qty is not None else None,
            item.get("unit"),
            item.get("notes"),
        ]
        for col, value in enumerate(values, start=1):
            # _formula_safe: descriptions and notes are free text lifted out of
            # the GC's BOQ — a leading '=' would otherwise land in the vendor's
            # workbook as a live formula (CWE-1236), same guard as the exports.
            c = ws.cell(row=r, column=col, value=_formula_safe(value))
            c.font = body_font
            c.border = border
            c.alignment = aligns[col - 1]
            if i % 2:
                c.fill = band_fill
        ws.cell(row=r, column=3).number_format = "#,##0.##"
        r += 1

    # ---- Footer -------------------------------------------------------
    note_row = r + 1
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=ncols)
    note = ws.cell(row=note_row, column=1, value=_FOOTER_NOTE)
    note.font = Font(size=9, italic=True, color=_MUTED)
    note.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
    ws.row_dimensions[note_row].height = 26

    sign_row = note_row + 1
    ws.merge_cells(start_row=sign_row, start_column=1, end_row=sign_row, end_column=ncols)
    sign = ws.cell(
        row=sign_row,
        column=1,
        value=f"G3 Electrical  ·  Office {email_branding.OFFICE_PHONE_DISPLAY}",
    )
    sign.font = Font(size=9, bold=True, color=_NAVY)
    sign.alignment = Alignment(horizontal="center", vertical="center")

    # ---- Print set-up: portrait letter, fit to width, repeating header ---
    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.print_title_rows = f"{header_row}:{header_row}"
    ws.print_area = f"A1:{last_col}{sign_row}"
    ws.page_margins.left = ws.page_margins.right = 0.4
    ws.page_margins.top = ws.page_margins.bottom = 0.5
    ws.freeze_panes = f"A{header_row + 1}"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def rows_for_preview(xlsx_bytes: bytes, max_rows: int = 500) -> list[list[Any]]:
    """Parse a stored .xlsx back into a list of rows for server-side preview."""
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True, read_only=True)
    rows: list[list[Any]] = []
    ws = wb.active
    for row in ws.iter_rows(values_only=True):
        vals = list(row)
        while vals and (vals[-1] is None or vals[-1] == ""):
            vals.pop()
        rows.append(["" if v is None else v for v in vals])
        if len(rows) >= max_rows:
            break
    wb.close()
    return rows
